"""
==============================================================
 MVO PORTFOLIO CONSTRUCTION & RISK ANALYSIS PIPELINE  v2.0
==============================================================
 MODOS DE ENTRADA:
   A) Excel / CSV  — importa precios históricos desde archivo
   B) yfinance     — descarga automática desde Yahoo Finance

 FLUJO:
   1  Adquisición de datos
   2  QC + retornos logarítmicos
   3  Análisis descriptivo y distribucional
   4  Correlación de Spearman + clustering jerárquico
   5  Optimización MVO (6 portafolios + frontera eficiente)
   6  Métricas de riesgo de mercado
   7  Exportación completa a Excel (un solo archivo)
   8  Visualizaciones PNG

 PORTAFOLIOS GENERADOS:
   MSR  — Máximo Sharpe Ratio
   GMV  — Mínima Varianza Global
   MCVR — Mínimo CVaR 95%
   MXR  — Máximo Retorno (dado β_p ≤ 1)
   RP   — Risk Parity (igual contribución al riesgo)
   EW   — Igual Ponderación (1/N, benchmark interno)
   FE   — N puntos sobre la frontera eficiente

 MÉTRICAS POR PORTAFOLIO:
   Sharpe, Sortino, Calmar, Treynor, Jensen α (t-stat + p)
   Beta, R², VaR 95/99 (hist + normal), CVaR 95/99
   Max Drawdown, Diversification Ratio, HHI, N_ef, MCTR/PCTR

 Autor : Fabiana Minaya Castillo
==============================================================
"""

# ──────────────────────────────────────────────────────────────
# 0. IMPORTS
# ──────────────────────────────────────────────────────────────
import os, sys, warnings
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from scipy import stats, optimize
from scipy.cluster import hierarchy
from scipy.spatial.distance import squareform
from statsmodels.stats.diagnostic import het_arch

import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.lines import Line2D

import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox

try:
    import yfinance as yf
    YF_OK = True
except ImportError:
    YF_OK = False

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    XL_OK = True
except ImportError:
    XL_OK = False

warnings.filterwarnings('ignore')

# ──────────────────────────────────────────────────────────────
# 1. PARÁMETROS GLOBALES  (edita aquí antes de ejecutar)
# ──────────────────────────────────────────────────────────────
RF_ANUAL    = 0.0525   # Tasa libre de riesgo anual (5.25 % SOFR ene-2025)
ANNUAL      = 252      # Días hábiles por año
N_FRONTIER  = 80       # Puntos en la frontera eficiente
WINDOW      = 30       # Ventana rolling para volatilidad (días)
ALPHAS      = [0.95, 0.99]
CAP_MAX     = 0.30     # Peso máximo por activo (0 < CAP ≤ 1 ; 1 = sin límite)
MIN_W       = 0.00     # Peso mínimo por activo (long-only)
N_INIT      = 150      # Inicializaciones Dirichlet para evitar mínimos locales
BENCHMARK   = 'SPY'   # Ticker del benchmark

# ──────────────────────────────────────────────────────────────
# 2. PALETA VISUAL
# ──────────────────────────────────────────────────────────────
BG    = '#0D1117'; PANEL = '#161B22'; TEXT  = '#E6EDF3'
MUTED = '#8B949E'; GRID  = '#21262D'
BLUE  = '#3B82F6'; RED   = '#EF4444'; GREEN = '#22C55E'
AMBER = '#F59E0B'; VIO   = '#7C3AED'; TEAL  = '#14B8A6'
COLS  = ['#3B82F6','#EF4444','#22C55E','#F59E0B','#7C3AED',
          '#EC4899','#14B8A6','#F97316','#A855F7','#6366F1',
          '#0EA5E9','#84CC16','#FB923C','#C084FC','#34D399']

plt.rcParams.update({
    'figure.facecolor': BG,   'axes.facecolor':  PANEL,
    'axes.edgecolor':  GRID,  'axes.labelcolor': TEXT,
    'xtick.color':    MUTED,  'ytick.color':     MUTED,
    'text.color':     TEXT,   'grid.color':      GRID,
    'grid.linewidth': 0.4,    'font.family':     'DejaVu Sans',
    'legend.facecolor': PANEL,'legend.edgecolor': MUTED,
    'legend.labelcolor': TEXT,'axes.grid': True,
    'grid.alpha': 0.35,       'axes.spines.top': False,
    'axes.spines.right': False,
})

# ──────────────────────────────────────────────────────────────
# 3. UTILIDADES GUI
# ──────────────────────────────────────────────────────────────
def _root():
    r = tk.Tk(); r.withdraw(); r.attributes('-topmost', True)
    return r

def gui_archivo(titulo, tipos=None):
    r = _root()
    tipos = tipos or [("Excel/CSV","*.xlsx *.xls *.csv"),("Todos","*.*")]
    ruta  = filedialog.askopenfilename(title=titulo, filetypes=tipos)
    r.destroy()
    if not ruta:
        print("No se seleccionó archivo. Saliendo."); sys.exit(1)
    return ruta

def gui_carpeta(titulo="Selecciona carpeta de salida"):
    r = _root()
    ruta = filedialog.askdirectory(title=titulo)
    r.destroy()
    if not ruta:
        print("No se seleccionó carpeta. Saliendo."); sys.exit(1)
    os.makedirs(ruta, exist_ok=True)
    return ruta

def gui_string(titulo, prompt, default=""):
    r = _root()
    val = simpledialog.askstring(title=titulo, prompt=prompt,
                                  initialvalue=default, parent=r)
    r.destroy()
    return val or default

def gui_float(titulo, prompt, default):
    r = _root()
    val = simpledialog.askfloat(title=titulo, prompt=prompt,
                                 initialvalue=default, parent=r)
    r.destroy()
    return val if val is not None else default

def gui_modo():
    """Diálogo para seleccionar modo A (Excel/CSV) o B (yfinance)."""
    r = _root()
    modo = tk.StringVar(value='A')
    win  = tk.Toplevel(r); win.title("Modo de entrada")
    win.attributes('-topmost', True)
    win.geometry("340x180"); win.resizable(False, False)
    tk.Label(win, text="Selecciona el modo de entrada de datos:",
             font=("Helvetica", 11, "bold"), pady=10).pack()
    tk.Radiobutton(win, text="A)  Importar Excel / CSV",
                   variable=modo, value='A', font=("Helvetica", 10)).pack(anchor='w', padx=40)
    tk.Radiobutton(win, text="B)  Descargar vía yfinance",
                   variable=modo, value='B',
                   font=("Helvetica", 10),
                   state='normal' if YF_OK else 'disabled').pack(anchor='w', padx=40)
    tk.Button(win, text="Continuar →",
              command=win.destroy, width=14).pack(pady=16)
    win.wait_window()
    r.destroy()
    return modo.get()

def _guardar(fig, path):
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor=BG)
    plt.show(); plt.close(fig)
    print(f"    ✓ {os.path.basename(path)}")
    return path

# ──────────────────────────────────────────────────────────────
# 4. ADQUISICIÓN DE DATOS
# ──────────────────────────────────────────────────────────────
def cargar_excel_csv(ruta: str) -> pd.DataFrame:
    """
    Carga precios desde Excel (.xlsx / .xls) o CSV (.csv).
    El archivo debe tener una columna de fechas (primera columna o 'Fecha')
    y una columna por activo con precios de cierre ajustados.
    """
    ext = os.path.splitext(ruta)[1].lower()
    if ext in ('.xlsx', '.xls'):
        df = pd.read_excel(ruta, index_col=0, parse_dates=True)
    elif ext == '.csv':
        df = pd.read_csv(ruta, index_col=0, parse_dates=True)
    else:
        raise ValueError(f"Formato no soportado: {ext}")
    df.index = pd.to_datetime(df.index)
    df = df.sort_index()
    df = df.apply(pd.to_numeric, errors='coerce')
    return df

def cargar_yfinance(tickers: list, start: str, end: str) -> pd.DataFrame:
    """
    Descarga precios de cierre ajustados desde Yahoo Finance.
    tickers : lista de strings (ej. ['AAPL','MSFT','LLY'])
    start/end: strings 'YYYY-MM-DD'
    """
    raw = yf.download(tickers, start=start, end=end,
                       auto_adjust=True, progress=True)
    if isinstance(raw.columns, pd.MultiIndex):
        df = raw['Close'][tickers]
    else:
        df = raw['Close'].to_frame(name=tickers[0])
    df.index = pd.to_datetime(df.index)
    df.columns.name = None
    return df.sort_index()

def descargar_benchmark(start: str, end: str, ticker: str = BENCHMARK):
    """Descarga retornos logarítmicos del benchmark."""
    if not YF_OK:
        return None
    try:
        raw = yf.download(ticker, start=start, end=end,
                           auto_adjust=True, progress=False)
        ret = np.log(raw['Close'] / raw['Close'].shift(1)).dropna()
        return ret.squeeze()
    except Exception as e:
        print(f"  ⚠ No se pudo descargar {ticker}: {e}")
        return None

# ──────────────────────────────────────────────────────────────
# 5. QC + RETORNOS
# ──────────────────────────────────────────────────────────────
def qc_y_retornos(precios: pd.DataFrame):
    """
    Control de calidad y cálculo de retornos logarítmicos.
    r_t = ln(P_t / P_{t-1})

    QC aplicado:
      • Forward-fill (máx. 3 sesiones consecutivas) para precios faltantes.
      • Eliminación de columnas con >20% de NAs tras ffill.
      • Aviso si alguna columna tiene precios ≤ 0.
    """
    print(f"\n  Activos originales : {precios.shape[1]}")
    print(f"  Observaciones      : {precios.shape[0]}")
    print(f"  Periodo            : {precios.index[0].date()} → {precios.index[-1].date()}")

    # Precios no positivos
    neg = (precios <= 0).sum().sum()
    if neg > 0:
        print(f"  ⚠ {neg} precio(s) ≤ 0 detectados → reemplazados por NaN")
        precios[precios <= 0] = np.nan

    # Forward-fill
    precios_c = precios.ffill(limit=3)

    # Eliminar columnas con >20% NA
    pct_na    = precios_c.isna().mean()
    eliminar  = pct_na[pct_na > 0.20].index.tolist()
    if eliminar:
        print(f"  ⚠ Columnas eliminadas (>20% NA): {eliminar}")
        precios_c = precios_c.drop(columns=eliminar)

    print(f"  Activos finales    : {precios_c.shape[1]}")
    print(f"  NAs residuales     : {precios_c.isna().sum().sum()}")

    ret = np.log(precios_c / precios_c.shift(1)).dropna()
    print(f"  Retornos log (N)   : {ret.shape[0]} obs × {ret.shape[1]} activos")
    return precios_c, ret

# ──────────────────────────────────────────────────────────────
# 6. ANÁLISIS DESCRIPTIVO Y DISTRIBUCIONAL
# ──────────────────────────────────────────────────────────────
def analisis_descriptivo(ret: pd.DataFrame) -> pd.DataFrame:
    """
    Estadísticos descriptivos + tests de normalidad por activo.

    Tests:
      JB  : Jarque-Bera  H0: normalidad  (chi² → global)
      SW  : Shapiro-Wilk H0: normalidad  (más potente con n < 5000)
      DA  : D'Agostino K² H0: normalidad (robusto a n grande)
    """
    rf_d  = (1 + RF_ANUAL) ** (1/ANNUAL) - 1
    desc  = pd.DataFrame(index=ret.columns)

    desc['μ_diario']    = ret.mean()
    desc['μ_anual_%']   = ret.mean() * ANNUAL * 100
    desc['σ_diario']    = ret.std()
    desc['σ_anual_%']   = ret.std() * np.sqrt(ANNUAL) * 100
    desc['Sharpe_ind']  = (ret.mean() - rf_d) / ret.std() * np.sqrt(ANNUAL)
    desc['Min_%']       = ret.min() * 100
    desc['Max_%']       = ret.max() * 100
    desc['Asimetría_G1'] = ret.skew()
    desc['Curtosis_G2'] = ret.kurt()

    jb_s, jb_p = [], []
    sw_s, sw_p = [], []
    da_s, da_p = [], []

    for col in ret.columns:
        s = ret[col].dropna().values
        j = stats.jarque_bera(s)
        w = stats.shapiro(s)
        d = stats.normaltest(s)
        jb_s.append(j.statistic); jb_p.append(j.pvalue)
        sw_s.append(w.statistic); sw_p.append(w.pvalue)
        da_s.append(d.statistic); da_p.append(d.pvalue)

    desc['JB_stat'] = jb_s; desc['JB_p']   = jb_p
    desc['SW_stat'] = sw_s; desc['SW_p']   = sw_p
    desc['DA_stat'] = da_s; desc['DA_p']   = da_p
    desc['Rechaza_H0_normal'] = (
        (desc['JB_p'] < 0.05) |
        (desc['SW_p'] < 0.05) |
        (desc['DA_p'] < 0.05)
    )

    # ARCH-LM (5 rezagos)
    arch_p = []
    for col in ret.columns:
        try:
            _, p, _, _ = het_arch(ret[col].dropna().values, nlags=5)
            arch_p.append(p)
        except Exception:
            arch_p.append(np.nan)
    desc['ARCH_LM_p'] = arch_p
    desc['ARCH_presente'] = desc['ARCH_LM_p'] < 0.05

    # VaR y CVaR histórico por activo
    for a in ALPHAS:
        lbl = int(a * 100)
        thr = ret.quantile(1 - a)
        desc[f'VaR_hist_{lbl}%'] = -thr * 100
        desc[f'CVaR_hist_{lbl}%'] = ret.apply(
            lambda c: -c[c <= thr[c.name]].mean() * 100
        )

    n_rej = desc['Rechaza_H0_normal'].sum()
    print(f"\n  {n_rej}/{len(desc)} activos rechazan normalidad al 5%")
    n_arch = desc['ARCH_presente'].sum()
    print(f"  {n_arch}/{len(desc)} activos con efectos ARCH (p<0.05, lag=5)")
    return desc

# ──────────────────────────────────────────────────────────────
# 7. CORRELACIÓN DE SPEARMAN
# ──────────────────────────────────────────────────────────────
def analisis_correlacion(ret: pd.DataFrame):
    """
    Matriz de correlación de Spearman + p-valores + clustering Ward.

    Spearman sobre Pearson porque:
      • Los retornos financieros violan normalidad (confirmado en Fase 3).
      • Spearman es robusto a outliers y relaciones no lineales.
    """
    n    = len(ret.columns)
    corr = ret.corr(method='spearman')
    pval = pd.DataFrame(np.ones((n, n)),
                         index=ret.columns, columns=ret.columns)
    for i in range(n):
        for j in range(n):
            if i != j:
                _, p = stats.spearmanr(ret.iloc[:, i].dropna(),
                                       ret.iloc[:, j].dropna())
                pval.iloc[i, j] = p

    dist_c    = squareform(np.clip(1 - np.abs(corr.values), 0, None))
    linkage   = hierarchy.linkage(dist_c, method='ward')
    ord_idx   = hierarchy.leaves_list(
                    hierarchy.optimal_leaf_ordering(linkage, dist_c))
    corr_ord  = corr.iloc[ord_idx, :].iloc[:, ord_idx]
    pval_ord  = pval.iloc[ord_idx, :].iloc[:, ord_idx]

    n_pares  = n*(n-1)//2
    tri_mask = np.triu_indices(n, k=1)
    n_sig    = (pval.values[tri_mask] < 0.05).sum()
    print(f"\n  Pares totales     : {n_pares}")
    print(f"  Pares sig. p<0.05 : {n_sig}")
    print(f"  ρ medio (sig.)    : {corr.values[tri_mask][pval.values[tri_mask] < 0.05].mean():.4f}")
    return corr, pval, corr_ord, pval_ord, linkage, ord_idx

# ──────────────────────────────────────────────────────────────
# 8. MOTOR MVO
# ──────────────────────────────────────────────────────────────
def _sigma_p(w, Sigma):
    """Volatilidad del portafolio."""
    return np.sqrt(w @ Sigma @ w)

def _mu_p(w, mu):
    """Retorno esperado del portafolio."""
    return w @ mu

def _sharpe(w, mu, Sigma, rf_d):
    """Sharpe ratio negativo (para minimizar)."""
    s = _sigma_p(w, Sigma)
    if s < 1e-10:
        return 0.0
    return -((_mu_p(w, mu) - rf_d) / s)

def _varianza(w, Sigma):
    return w @ Sigma @ w

def _rp_obj(w, Sigma):
    """
    Risk Parity: minimizar Σᵢ(MCTR_i − σ_p/N)²
    Cada activo contribuye igualmente al riesgo total.
    """
    s    = _sigma_p(w, Sigma)
    if s < 1e-10: return 0.0
    Sw   = Sigma @ w
    mctr = w * Sw / s
    target = s / len(w)
    return np.sum((mctr - target) ** 2)

def _cvar_obj(w, ret_matrix, alpha=0.95):
    """
    CVaR histórico del portafolio (para minimizar).
    ret_matrix: np.array (T × N)
    """
    port_r = ret_matrix @ w
    thresh = np.percentile(port_r, (1 - alpha) * 100)
    tail   = port_r[port_r <= thresh]
    return -tail.mean() if len(tail) > 0 else 0.0

def _max_ret_obj(w, mu):
    return -_mu_p(w, mu)

def _constraints_base(n: int):
    """Restricciones comunes: suma de pesos = 1."""
    return [{'type': 'eq', 'fun': lambda w: np.sum(w) - 1.0}]

def _bounds(n: int):
    return [(MIN_W, CAP_MAX)] * n

def _slsqp(objective, n, constraints, extra_kwargs=None):
    """
    Minimiza `objective` mediante SLSQP con N_INIT inicializaciones Dirichlet.
    Devuelve el resultado factible con menor valor de objetivo.
    """
    best_res, best_val = None, np.inf
    bnds = _bounds(n)
    rng  = np.random.default_rng(42)

    for _ in range(N_INIT):
        w0  = rng.dirichlet(np.ones(n))
        w0  = np.clip(w0, MIN_W, CAP_MAX)
        w0 /= w0.sum()
        kwargs = {'x0': w0, 'method': 'SLSQP',
                  'bounds': bnds, 'constraints': constraints,
                  'options': {'maxiter': 1000, 'ftol': 1e-9}}
        if extra_kwargs:
            kwargs.update(extra_kwargs)
        try:
            res = optimize.minimize(objective, **kwargs)
            if res.success and res.fun < best_val:
                best_val = res.fun
                best_res = res
        except Exception:
            continue

    if best_res is None:
        # fallback: igual ponderación
        class _Fallback:
            x = np.ones(n) / n
            success = False
        return _Fallback()
    # Limpiar pesos pequeños
    w = best_res.x.copy()
    w[w < 0.005] = 0.0
    if w.sum() > 0:
        w /= w.sum()
    best_res.x = w
    return best_res

def optimizar_portafolios(mu: np.ndarray, Sigma: np.ndarray,
                           rf_d: float, ret_matrix: np.ndarray,
                           tickers: list):
    """
    Genera 6 portafolios óptimos.

    Parámetros:
      mu         : retornos esperados diarios (N,)
      Sigma      : matriz de covarianza diaria (N×N)
      rf_d       : tasa libre de riesgo diaria
      ret_matrix : matriz de retornos históricos (T×N)
      tickers    : lista de nombres de activos

    Devuelve dict {nombre: np.ndarray de pesos}
    """
    n   = len(tickers)
    con = _constraints_base(n)

    portafolios = {}

    # ── MSR: Máximo Sharpe Ratio ──────────────────────────────
    print("    Optimizando MSR (Máximo Sharpe)...")
    res = _slsqp(lambda w: _sharpe(w, mu, Sigma, rf_d), n, con)
    portafolios['MSR'] = res.x

    # ── GMV: Mínima Varianza Global ───────────────────────────
    print("    Optimizando GMV (Mínima Varianza)...")
    res = _slsqp(lambda w: _varianza(w, Sigma), n, con)
    portafolios['GMV'] = res.x

    # ── RP: Risk Parity ───────────────────────────────────────
    print("    Optimizando RP (Risk Parity)...")
    res = _slsqp(lambda w: _rp_obj(w, Sigma), n, con)
    portafolios['RP'] = res.x

    # ── MCVR: Mínimo CVaR 95% ────────────────────────────────
    print("    Optimizando MCVR (Mínimo CVaR 95%)...")
    res = _slsqp(lambda w: _cvar_obj(w, ret_matrix, 0.95), n, con)
    portafolios['MCVR'] = res.x

    # ── MXR: Máximo Retorno con β_p ≤ 1  ─────────────────────
    # Si no hay benchmark, simplemente maximiza retorno sin restricción de beta.
    print("    Optimizando MXR (Máximo Retorno)...")
    con_mxr = con.copy()
    # Restricción Sharpe ≥ 1.5 para evitar portafolios triviales
    sr_floor = lambda w: (-_sharpe(w, mu, Sigma, rf_d) - (-1.5))
    con_mxr_ext = con + [{'type': 'ineq', 'fun': sr_floor}]
    res = _slsqp(lambda w: _max_ret_obj(w, mu), n, con_mxr_ext)
    portafolios['MXR'] = res.x if res.success else np.ones(n)/n

    # ── EW: Igual Ponderación (benchmark interno) ─────────────
    portafolios['EW'] = np.ones(n) / n

    return portafolios

def frontera_eficiente(mu: np.ndarray, Sigma: np.ndarray,
                        rf_d: float, n_pts: int = N_FRONTIER):
    """
    Traza la frontera eficiente resolviendo:
      min wᵀΣw  s.t. wᵀμ ≥ μ*, Σwᵢ=1, wᵢ≥0
    para N_FRONTIER valores de μ* entre μ_GMV y μ_max.

    Devuelve (vols, rets, pesos) — arrays de forma (n_pts,) y (n_pts, N)
    """
    n   = len(mu)
    con = _constraints_base(n)
    bnds = _bounds(n)
    rng  = np.random.default_rng(99)

    # Determinar rango de retornos factibles
    # GMV: retorno mínimo
    res_gmv = _slsqp(lambda w: _varianza(w, Sigma), n, con)
    mu_min  = _mu_p(res_gmv.x, mu)
    mu_max  = mu.max()

    targets  = np.linspace(mu_min, mu_max * 0.99, n_pts)
    vols, rets, pesos = [], [], []

    for mu_t in targets:
        con_t = con + [{'type': 'ineq',
                         'fun': lambda w, mt=mu_t: _mu_p(w, mu) - mt}]
        best_res, best_val = None, np.inf
        for _ in range(30):  # pocas inicializaciones para velocidad
            w0  = rng.dirichlet(np.ones(n))
            w0  = np.clip(w0, MIN_W, CAP_MAX); w0 /= w0.sum()
            try:
                res = optimize.minimize(
                    lambda w: _varianza(w, Sigma), w0,
                    method='SLSQP', bounds=bnds, constraints=con_t,
                    options={'maxiter': 500, 'ftol': 1e-9})
                if res.success and res.fun < best_val:
                    best_val = res.fun; best_res = res
            except Exception:
                continue
        if best_res is not None:
            w = best_res.x.copy()
            w[w < 0.005] = 0.0
            if w.sum() > 0:
                w /= w.sum()
            vols.append(_sigma_p(w, Sigma) * np.sqrt(ANNUAL) * 100)
            rets.append(_mu_p(w, mu) * ANNUAL * 100)
            pesos.append(w)

    return np.array(vols), np.array(rets), pesos

# ──────────────────────────────────────────────────────────────
# 9. MCTR / PCTR
# ──────────────────────────────────────────────────────────────
def calcular_mctr(w: np.ndarray, Sigma: np.ndarray, tickers: list):
    """
    Contribución Marginal al Riesgo (MCTR) y Porcentual (PCTR).

    MCTR_i = w_i × (Σw)_i / σ_p
    PCTR_i = MCTR_i / σ_p           (suma = 1)

    Interpretación:
      PCTR_i > w_i  → activo concentra riesgo por encima de su peso
      PCTR_i < w_i  → activo diversifica (aporta menos riesgo que peso)
    """
    s    = _sigma_p(w, Sigma)
    if s < 1e-12:
        return pd.Series(0.0, index=tickers), pd.Series(1/len(tickers), index=tickers)
    Sw   = Sigma @ w
    mctr = w * Sw / s
    pctr = mctr / s
    return (pd.Series(mctr, index=tickers),
            pd.Series(pctr / pctr.sum(), index=tickers))

# ──────────────────────────────────────────────────────────────
# 10. MÉTRICAS DE RIESGO POR PORTAFOLIO
# ──────────────────────────────────────────────────────────────
def calcular_metricas(port_ret: pd.Series,
                       mkt_ret:  pd.Series | None,
                       rf_d:     float) -> dict:
    """
    Calcula el conjunto completo de métricas de riesgo de mercado
    para una serie de retornos diarios de portafolio.

    Parámetros:
      port_ret : retornos log diarios del portafolio
      mkt_ret  : retornos log diarios del benchmark (puede ser None)
      rf_d     : tasa libre de riesgo diaria

    Métricas calculadas:
      ─ Retorno/riesgo básicos ─────────────────────────────────
      ret_anual    : retorno medio anualizado (%)
      vol_anual    : volatilidad anualizada (%)
      ret_acum     : retorno acumulado del período (%)

      ─ Ratios ajustados por riesgo ────────────────────────────
      sharpe       : (μ_p − rf) / σ_p  × √252
      sortino      : (μ_p − rf) / σ_dd × √252
                     σ_dd = downside deviation (retornos < 0)
      calmar       : ret_anual / |MDD|
      treynor      : (μ_p − rf) / β × √252   [NaN si no hay mkt]
      info_ratio   : jensen_alpha / tracking_error  [NaN si no hay mkt]

      ─ Riesgo sistemático (requiere benchmark) ────────────────
      beta         : Cov(r_p, r_m) / Var(r_m)
      jensen_alpha : α = R_p − [rf + β(R_m − rf)]  (anualizado %)
      alpha_pval   : p-valor de α vía OLS
      r2           : R² de la regresión CAPM
      riesgo_idio  : σ²_p − β²σ²_m   (varianza idiosincrática)

      ─ Riesgo de cola ─────────────────────────────────────────
      var_hist_95/99  : VaR histórico (percentil) diario (%)
      var_norm_95/99  : VaR paramétrico normal diario (%)
      cvar_95/99      : CVaR/ES histórico diario (%)

      ─ Drawdown ───────────────────────────────────────────────
      mdd          : Maximum Drawdown (%)
      mdd_dur      : duración del mayor drawdown (días)

      ─ Construcción ───────────────────────────────────────────
      div_ratio    : calculado externamente (ver bloque de portafolios)
    """
    m    = {}
    r    = port_ret.dropna().values
    rf_a = RF_ANUAL

    # ─ Básicos ─────────────────────────────────────────────────
    m['ret_anual_%']  = r.mean() * ANNUAL * 100
    m['vol_anual_%']  = r.std() * np.sqrt(ANNUAL) * 100
    m['ret_acum_%']   = (np.expm1(np.sum(r))) * 100

    # ─ Sharpe ──────────────────────────────────────────────────
    exc              = r - rf_d
    m['sharpe']      = exc.mean() / r.std() * np.sqrt(ANNUAL) if r.std() > 0 else np.nan

    # ─ Sortino ─────────────────────────────────────────────────
    down             = r[r < 0]
    sigma_dd         = down.std() if len(down) > 1 else np.nan
    m['sortino']     = exc.mean() / sigma_dd * np.sqrt(ANNUAL) if sigma_dd and sigma_dd > 0 else np.nan

    # ─ Maximum Drawdown ─────────────────────────────────────────
    cum              = np.exp(np.cumsum(r))
    roll_max         = np.maximum.accumulate(cum)
    dd               = (cum - roll_max) / roll_max
    m['mdd_%']       = dd.min() * 100
    # Duración del mayor drawdown
    below_zero       = dd < -1e-6
    if below_zero.any():
        starts = np.where(np.diff(below_zero.astype(int)) == 1)[0]
        ends   = np.where(np.diff(below_zero.astype(int)) == -1)[0]
        if len(starts) > 0 and len(ends) > 0:
            ends = ends[ends > starts[0]]
            if len(ends) > 0:
                durs = ends[:len(starts)] - starts[:len(ends)]
                m['mdd_dur_dias'] = int(durs.max()) if len(durs) > 0 else 0
            else:
                m['mdd_dur_dias'] = int(len(r) - starts[0])
        else:
            m['mdd_dur_dias'] = 0
    else:
        m['mdd_dur_dias'] = 0

    # ─ Calmar ──────────────────────────────────────────────────
    m['calmar']      = m['ret_anual_%'] / abs(m['mdd_%']) if m['mdd_%'] != 0 else np.nan

    # ─ VaR y CVaR ──────────────────────────────────────────────
    for a in ALPHAS:
        lbl = int(a * 100)
        thr = np.percentile(r, (1 - a) * 100)
        m[f'var_hist_{lbl}%']  = -thr * 100
        tail = r[r <= thr]
        m[f'cvar_hist_{lbl}%'] = -tail.mean() * 100 if len(tail) > 0 else np.nan
        m[f'var_norm_{lbl}%']  = -(r.mean() + stats.norm.ppf(1-a) * r.std()) * 100

    # ─ Riesgo sistemático ──────────────────────────────────────
    if mkt_ret is not None:
        common = port_ret.index.intersection(mkt_ret.index)
        rp_c   = port_ret.loc[common].values
        rm_c   = mkt_ret.loc[common].values
        if len(rp_c) > 10:
            cov_pm  = np.cov(rp_c, rm_c)
            beta    = cov_pm[0, 1] / cov_pm[1, 1]
            alpha_d = rp_c.mean() - (rf_d + beta * (rm_c.mean() - rf_d))
            # OLS para p-valor de alpha
            X       = np.column_stack([np.ones(len(rm_c)), rm_c])
            try:
                b, res_ols, _, _ = np.linalg.lstsq(X, rp_c, rcond=None)
                s2       = res_ols[0] / (len(rp_c) - 2) if len(res_ols) > 0 else np.var(rp_c - X@b)
                XtX_inv  = np.linalg.inv(X.T @ X)
                se_alpha = np.sqrt(s2 * XtX_inv[0, 0])
                t_alpha  = b[0] / se_alpha
                pval_a   = 2 * stats.t.sf(abs(t_alpha), df=len(rp_c)-2)
            except Exception:
                pval_a = np.nan
            r2       = np.corrcoef(rp_c, rm_c)[0,1]**2
            te       = (rp_c - rm_c).std() * np.sqrt(ANNUAL) * 100
            m['beta']          = beta
            m['jensen_alpha_%']= alpha_d * ANNUAL * 100
            m['alpha_pval']    = pval_a
            m['r2']            = r2
            m['riesgo_sist_%'] = (beta**2 * np.var(rm_c)) / np.var(rp_c) * 100
            m['riesgo_idio_%'] = 100 - m['riesgo_sist_%']
            m['treynor']       = exc.mean() / beta * np.sqrt(ANNUAL) if beta != 0 else np.nan
            m['tracking_err_%']= te
            m['info_ratio']    = (m['jensen_alpha_%'] / te) if te > 0 else np.nan
    else:
        for k in ['beta','jensen_alpha_%','alpha_pval','r2',
                  'riesgo_sist_%','riesgo_idio_%','treynor',
                  'tracking_err_%','info_ratio']:
            m[k] = np.nan

    return m

# ──────────────────────────────────────────────────────────────
# 11. VISUALIZACIONES
# ──────────────────────────────────────────────────────────────
def grafico_distribuciones(ret: pd.DataFrame, desc: pd.DataFrame,
                            out_dir: str, base: str):
    tickers = list(ret.columns)
    n       = len(tickers)
    ncols   = min(5, n)
    nrows   = (n + ncols - 1) // ncols

    fig, axes = plt.subplots(nrows, ncols,
                              figsize=(ncols*5, nrows*3.5),
                              facecolor=BG,
                              gridspec_kw={'hspace': 0.5, 'wspace': 0.3})
    fig.suptitle('Distribución Empírica de Retornos Logarítmicos  ·  '
                 f'KDE vs Normal teórica  ·  n={ret.shape[0]} obs',
                 fontsize=13, fontweight='bold', color=TEXT, y=1.01)

    flat = np.array(axes).flatten()
    for idx, tk in enumerate(tickers):
        ax  = flat[idx]
        r   = ret[tk].dropna().values * 100
        col = COLS[idx % len(COLS)]
        ax.hist(r, bins=min(60, max(30, len(r)//20)),
                density=True, color=col, alpha=0.35, edgecolor='none')
        xg  = np.linspace(r.min(), r.max(), 400)
        ax.plot(xg, stats.gaussian_kde(r)(xg), color=col, lw=1.8)
        ax.plot(xg, stats.norm.pdf(xg, r.mean(), r.std()),
                color=TEXT, lw=1.2, ls='--', alpha=0.55)
        rej = desc.loc[tk, 'Rechaza_H0_normal'] if tk in desc.index else False
        ax.set_title(tk, fontsize=9, fontweight='bold',
                     color=RED if rej else GREEN)
        jb_p = desc.loc[tk, 'JB_p'] if tk in desc.index else np.nan
        pstr = f'{jb_p:.2e}' if jb_p < 0.001 else f'{jb_p:.4f}'
        ax.text(0.97, 0.96,
                f"G₁={desc.loc[tk,'Asimetría_G1']:.2f}\n"
                f"G₂={desc.loc[tk,'Curtosis_G2']:.1f}\n"
                f"JB p={pstr}",
                transform=ax.transAxes, va='top', ha='right',
                fontsize=7, color=TEXT,
                bbox=dict(boxstyle='round,pad=0.3', fc=BG,
                          ec=GRID, alpha=0.85))
        ax.set_xlabel('Retorno diario (%)', fontsize=7)
        ax.set_ylabel('Densidad', fontsize=7)
        ax.tick_params(labelsize=6.5)

    for j in range(idx + 1, len(flat)):
        flat[j].set_visible(False)

    fig.legend(handles=[
        Line2D([0],[0], color=COLS[0], lw=2, label='KDE empírica'),
        Line2D([0],[0], color=TEXT, lw=1.2, ls='--', alpha=0.6,
               label='Normal N(μ,σ²)'),
        Line2D([0],[0], color=RED, lw=0, marker='s', ms=8,
               label='Rojo = rechaza normalidad'),
    ], loc='lower center', ncol=3,
    bbox_to_anchor=(0.5, -0.02), fontsize=9)

    path = os.path.join(out_dir, f'{base}_distribuciones.png')
    return _guardar(fig, path)


def grafico_correlacion(corr_ord: pd.DataFrame,
                         pval_ord: pd.DataFrame,
                         out_dir: str, base: str):
    n   = len(corr_ord)
    tks = list(corr_ord.columns)
    cmap = LinearSegmentedColormap.from_list(
        'sp', [(0,'#2563EB'),(0.4,'#93C5FD'),(0.5,'#F8FAFC'),
               (0.6,'#FCA5A5'),(1,'#DC2626')])

    fig, ax = plt.subplots(figsize=(max(8, n*0.9), max(7, n*0.8)),
                            facecolor=BG)
    im = ax.imshow(corr_ord.values, cmap=cmap, vmin=-1, vmax=1,
                   aspect='auto', interpolation='none')

    for i in range(n):
        for j in range(n):
            val = corr_ord.values[i, j]
            p   = pval_ord.values[i, j]
            if i == j:
                ax.text(j, i, '1.00', ha='center', va='center',
                        fontsize=max(5, 9-n//5), color='#1E293B')
            elif p >= 0.05:
                ax.text(j, i, '—', ha='center', va='center',
                        fontsize=max(5, 9-n//5), color=MUTED, alpha=0.5)
            else:
                tc = '#1E293B' if abs(val) < 0.45 else TEXT
                ax.text(j, i, f'{val:.2f}', ha='center', va='center',
                        fontsize=max(5, 9-n//5), color=tc)

    ax.set_xticks(range(n)); ax.set_yticks(range(n))
    ax.set_xticklabels(tks, rotation=45, ha='right',
                        fontsize=max(7, 10-n//8))
    ax.set_yticklabels(tks, fontsize=max(7, 10-n//8))
    cbar = fig.colorbar(im, ax=ax, fraction=0.03, pad=0.02)
    cbar.set_label('ρ Spearman', color=TEXT, fontsize=9)
    plt.setp(cbar.ax.yaxis.get_ticklabels(), color=TEXT)
    ax.set_title('Correlación de Spearman  ·  Clustering Ward  ·  '
                 '"—" = p ≥ 0.05',
                 fontsize=11, fontweight='bold', color=TEXT, pad=12)
    plt.tight_layout()
    path = os.path.join(out_dir, f'{base}_correlacion.png')
    return _guardar(fig, path)


def grafico_frontera(fe_vols: np.ndarray, fe_rets: np.ndarray,
                      portafolios_stats: dict,
                      out_dir: str, base: str):
    """Frontera eficiente con portafolios identificados."""
    PORT_COLORS = {
        'MSR': AMBER, 'GMV': GREEN, 'RP': VIO,
        'MCVR': TEAL, 'MXR': RED,  'EW': BLUE
    }
    PORT_MARKS  = {'MSR':'*','GMV':'D','RP':'P','MCVR':'v','MXR':'^','EW':'s'}

    fig, ax = plt.subplots(figsize=(11, 7), facecolor=BG)
    if len(fe_vols) > 1:
        ax.plot(fe_vols, fe_rets, color=MUTED, lw=1.5, ls='--',
                alpha=0.6, zorder=2, label='Frontera eficiente')

    for pname, stats_p in portafolios_stats.items():
        v = stats_p.get('vol_anual_%', np.nan)
        r = stats_p.get('ret_anual_%', np.nan)
        s = stats_p.get('sharpe', np.nan)
        if np.isnan(v) or np.isnan(r):
            continue
        col = PORT_COLORS.get(pname, COLS[0])
        mk  = PORT_MARKS.get(pname, 'o')
        ms  = 200 if pname == 'MSR' else 100
        ax.scatter(v, r, c=col, s=ms, marker=mk, zorder=5,
                   edgecolors='none',
                   label=f"{pname}  (Sharpe={s:.2f})" if not np.isnan(s) else pname)
        ax.annotate(pname, (v, r),
                    xytext=(v+0.3, r+0.3),
                    fontsize=8.5, color=col, fontweight='bold')

    # Línea CML desde RF
    rf_a = RF_ANUAL * 100
    if len(fe_vols) > 0:
        slope = (fe_rets.max() - rf_a) / fe_vols[np.argmax(fe_rets)]
        x_cml = np.linspace(0, fe_vols.max() * 1.1, 50)
        ax.plot(x_cml, rf_a + slope * x_cml,
                color=AMBER, lw=1.0, ls=':', alpha=0.5,
                label=f'CML (aprox. desde RF={rf_a:.2f}%)')

    ax.set_xlabel('Volatilidad Anualizada (%)', fontsize=11)
    ax.set_ylabel('Retorno Anualizado (%)',      fontsize=11)
    ax.set_title('Frontera Eficiente  ·  Portafolios Óptimos\n'
                 'Markowitz MVO  ·  Long-only  ·  '
                 f'Cap máx={CAP_MAX*100:.0f}%',
                 fontsize=12, fontweight='bold', color=TEXT, pad=10)
    ax.legend(fontsize=8.5, loc='lower right')
    ax.set_xlim(left=0)
    plt.tight_layout()
    path = os.path.join(out_dir, f'{base}_frontera.png')
    return _guardar(fig, path)


def grafico_pesos(weights: dict, tickers: list,
                   out_dir: str, base: str):
    """Gráfico de pesos por portafolio (stacked bar)."""
    port_names = list(weights.keys())
    n_p = len(port_names)
    n_t = len(tickers)
    mat = np.array([weights[p] for p in port_names])  # (n_p, n_t)

    fig, ax = plt.subplots(figsize=(max(10, n_p*1.2), 6), facecolor=BG)
    bottom  = np.zeros(n_p)
    x       = np.arange(n_p)

    for i, tk in enumerate(tickers):
        vals = mat[:, i]
        ax.bar(x, vals * 100, bottom=bottom, color=COLS[i % len(COLS)],
               alpha=0.85, label=tk if vals.max() > 0.01 else '_',
               width=0.65)
        for j, (v, b) in enumerate(zip(vals, bottom)):
            if v > 0.04:
                ax.text(j, b*100 + v*50, f'{v*100:.0f}%',
                        ha='center', va='center',
                        fontsize=max(6.5, 8-n_t//6),
                        color='#1E293B', fontweight='bold')
        bottom += vals

    ax.set_xticks(x); ax.set_xticklabels(port_names, fontsize=10,
                                           fontweight='bold')
    ax.set_ylabel('Peso (%)', fontsize=10)
    ax.set_ylim(0, 105)
    ax.set_title('Pesos por Portafolio  ·  MVO  ·  Long-only',
                 fontsize=12, fontweight='bold', color=TEXT, pad=10)
    ax.legend(bbox_to_anchor=(1.01, 1), loc='upper left',
              fontsize=8, ncol=max(1, n_t//15))
    ax.grid(axis='y', alpha=0.35)
    plt.tight_layout()
    path = os.path.join(out_dir, f'{base}_pesos.png')
    return _guardar(fig, path)


def grafico_mctr(mctr_dict: dict, pctr_dict: dict,
                  tickers: list, out_dir: str, base: str):
    """Peso vs. PCTR para cada portafolio."""
    port_names = list(mctr_dict.keys())
    n_p = len(port_names)
    ncols = min(3, n_p); nrows = (n_p + ncols - 1) // ncols

    fig, axes = plt.subplots(nrows, ncols,
                              figsize=(ncols*5, nrows*4), facecolor=BG,
                              gridspec_kw={'hspace': 0.5, 'wspace': 0.35})
    flat = np.array(axes).flatten()

    for idx, pname in enumerate(port_names):
        ax    = flat[idx]
        w     = pd.Series(mctr_dict[pname][0].values,
                           index=tickers)  # usando pesos directamente
        pctr  = pctr_dict[pname]
        active = [t for t in tickers if w.get(t, 0) > 0.005 or pctr.get(t, 0) > 0.005]
        if not active:
            ax.set_visible(False); continue

        x   = np.arange(len(active))
        w_a = np.array([w.get(t, 0) * 100 for t in active])
        p_a = np.array([pctr.get(t, 0) * 100 for t in active])
        col_bars = [GREEN if p_a[i] <= w_a[i] else RED
                    for i in range(len(active))]

        ax.barh(x - 0.2, w_a,  height=0.35, color=BLUE, alpha=0.8, label='Peso %')
        ax.barh(x + 0.2, p_a,  height=0.35, color=col_bars, alpha=0.75, label='PCTR %')
        ax.set_yticks(x); ax.set_yticklabels(active, fontsize=8)
        ax.set_xlabel('%', fontsize=8)
        ax.set_title(pname, fontsize=10, fontweight='bold', pad=5)
        ax.grid(axis='x', alpha=0.35)
        if idx == 0:
            ax.legend(fontsize=7)

    for j in range(idx+1, len(flat)):
        flat[j].set_visible(False)

    fig.suptitle('Peso vs. Contribución al Riesgo (PCTR)\n'
                 'Rojo = activo concentra riesgo por encima de su peso',
                 fontsize=12, fontweight='bold', color=TEXT, y=1.01)
    plt.tight_layout()
    path = os.path.join(out_dir, f'{base}_mctr.png')
    return _guardar(fig, path)


def grafico_metricas_resumen(metricas: dict, out_dir: str, base: str):
    """Heatmap de métricas clave por portafolio."""
    port_names = list(metricas.keys())
    keys_show  = ['ret_anual_%', 'vol_anual_%', 'sharpe', 'sortino',
                  'calmar', 'mdd_%', 'var_hist_95%', 'cvar_hist_95%',
                  'beta', 'jensen_alpha_%', 'diversif_ratio', 'hhi_%']
    keys_label = ['Ret. Anual %', 'Vol. Anual %', 'Sharpe', 'Sortino',
                  'Calmar', 'Max DD %', 'VaR 95% hist', 'CVaR 95% hist',
                  'Beta', 'Jensen α %', 'Div. Ratio', 'HHI %']

    mat = []
    for pk in keys_show:
        row = []
        for pn in port_names:
            val = metricas[pn].get(pk, np.nan)
            row.append(val if not np.isnan(val) else 0)
        mat.append(row)
    mat = np.array(mat, dtype=float)

    # Normalizar por fila para color relativo
    norm = np.zeros_like(mat)
    for i in range(len(mat)):
        rmin, rmax = mat[i].min(), mat[i].max()
        if rmax > rmin:
            norm[i] = (mat[i] - rmin) / (rmax - rmin)
        else:
            norm[i] = 0.5

    cmap = LinearSegmentedColormap.from_list(
        'rm', [(0,'#1E3A5F'),(0.5,'#F59E0B'),(1,'#EF4444')])

    fig, ax = plt.subplots(figsize=(max(10, len(port_names)*1.6),
                                     max(7, len(keys_show)*0.75)),
                            facecolor=BG)
    im = ax.imshow(norm, cmap=cmap, aspect='auto', vmin=0, vmax=1)

    for i in range(len(keys_show)):
        for j in range(len(port_names)):
            val = mat[i, j]
            fmt = f'{val:.2f}' if abs(val) < 100 else f'{val:.1f}'
            tc  = '#1E293B' if norm[i, j] < 0.55 else TEXT
            ax.text(j, i, fmt, ha='center', va='center',
                    fontsize=9, color=tc)

    ax.set_xticks(range(len(port_names)))
    ax.set_yticks(range(len(keys_show)))
    ax.set_xticklabels(port_names, fontsize=10, fontweight='bold')
    ax.set_yticklabels(keys_label, fontsize=9)
    ax.set_title('Resumen de Métricas  ·  Escala de color relativa por fila\n'
                 'Rojo = mayor valor relativo  |  Azul = menor',
                 fontsize=11, fontweight='bold', color=TEXT, pad=12)
    cbar = fig.colorbar(im, ax=ax, fraction=0.025, pad=0.02)
    cbar.set_label('Valor normalizado', color=TEXT, fontsize=8)
    plt.setp(cbar.ax.yaxis.get_ticklabels(), color=TEXT)
    plt.tight_layout()
    path = os.path.join(out_dir, f'{base}_resumen_metricas.png')
    return _guardar(fig, path)


def grafico_drawdown(port_returns: pd.DataFrame, port_names: list,
                      out_dir: str, base: str):
    """Drawdown acumulado por portafolio."""
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 9), facecolor=BG,
                                    gridspec_kw={'height_ratios':[3,1],
                                                 'hspace':0.3})
    max_dd = {}
    for i, pn in enumerate(port_names):
        r   = port_returns[pn].values
        cum = np.exp(np.cumsum(r))
        rm  = np.maximum.accumulate(cum)
        dd  = (cum - rm) / rm * 100
        max_dd[pn] = dd.min()
        col = AMBER if pn == 'MSR' else COLS[i % len(COLS)]
        lw  = 2.0 if pn == 'MSR' else 1.0
        ax1.plot(port_returns.index, dd, color=col, lw=lw,
                 alpha=0.9 if pn=='MSR' else 0.65, label=pn)

    ax1.set_ylabel('Drawdown (%)', fontsize=10)
    ax1.set_title('Drawdown Acumulado  ·  Todos los Portafolios  '
                  '·  ⭐ MSR en ámbar',
                  fontsize=11, fontweight='bold', color=TEXT, pad=8)
    ax1.legend(ncol=6, fontsize=8, loc='lower left')
    ax1.set_ylim(top=5)
    ax1.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
    ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=30, ha='right')

    cols_dd = [AMBER if p=='MSR' else COLS[i%len(COLS)]
               for i,p in enumerate(port_names)]
    ax2.bar(port_names, [max_dd[p] for p in port_names],
            color=cols_dd, alpha=0.85, width=0.65)
    for i, pn in enumerate(port_names):
        ax2.text(i, max_dd[pn]-0.3, f'{max_dd[pn]:.1f}%',
                 ha='center', va='top', fontsize=8, color=TEXT)
    ax2.set_ylabel('Max DD (%)', fontsize=9)
    ax2.set_title('Máximo Drawdown por Portafolio', fontsize=10, pad=5)

    plt.tight_layout()
    path = os.path.join(out_dir, f'{base}_drawdown.png')
    return _guardar(fig, path)

# ──────────────────────────────────────────────────────────────
# 12. EXPORTACIÓN A EXCEL
# ──────────────────────────────────────────────────────────────
def exportar_excel(precios, ret, desc, corr_sp, fe_vols, fe_rets,
                    weights, metricas, mctr_d, pctr_d,
                    tickers, port_names, out_dir, base,
                    img_paths: list):
    """
    Genera un único archivo Excel con las siguientes hojas:
      1  Resumen          — dashboard ejecutivo
      2  Precios          — precios de cierre ajustados
      3  Retornos_Log     — retornos logarítmicos diarios
      4  Descriptivos     — estadísticos + tests de normalidad
      5  Correlacion      — matriz Spearman
      6  Frontera         — puntos de la frontera eficiente
      7  Pesos            — pesos por portafolio
      8  Metricas         — todas las métricas de riesgo
      9  VaR_CVaR         — desglose VaR/CVaR por nivel
      10 MCTR_PCTR        — contribución al riesgo por activo
    """
    if not XL_OK:
        print("  ⚠ openpyxl no instalado. Excel omitido.")
        return

    path_xl = os.path.join(out_dir, f'{base}_analisis_portafolio.xlsx')

    # ── Estilos ──────────────────────────────────────────────
    hdr_fill  = PatternFill('solid', fgColor='264653')
    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    body_font = Font(name='Calibri', size=9)
    ca        = Alignment(horizontal='center', vertical='center')
    ra        = Alignment(horizontal='right',  vertical='center')
    la        = Alignment(horizontal='left',   vertical='center')
    thin      = Side(style='thin', color='DEE2E6')
    bord      = Border(bottom=thin)

    pos_fill  = PatternFill('solid', fgColor='D4EDDA')
    pos_font  = Font(name='Calibri', bold=True, color='155724', size=9)
    neg_fill  = PatternFill('solid', fgColor='F8D7DA')
    neg_font  = Font(name='Calibri', bold=True, color='721C24', size=9)
    warn_fill = PatternFill('solid', fgColor='FFF3CD')
    warn_font = Font(name='Calibri', bold=True, color='856404', size=9)

    def _fmt_hdr(ws):
        for cell in ws[1]:
            cell.fill  = hdr_fill
            cell.font  = hdr_font
            cell.alignment = ca
        ws.freeze_panes = 'B2'

    def _aw(ws):
        for col in ws.columns:
            ml = max((len(str(c.value)) for c in col
                      if c.value is not None), default=8)
            ws.column_dimensions[
                get_column_letter(col[0].column)].width = min(ml + 3, 30)

    def _write_df(ws, df, index=True, num_fmt='0.0000'):
        if index:
            df = df.reset_index()
        for r_idx, row in enumerate(
                [list(df.columns)] + df.values.tolist(), start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.fill = hdr_fill; cell.font = hdr_font
                    cell.alignment = ca
                else:
                    cell.font = body_font; cell.border = bord
                    if isinstance(val, float):
                        cell.number_format = num_fmt
                        cell.alignment = ra
                    elif isinstance(val, (int, np.integer)):
                        cell.alignment = ra
                    else:
                        cell.alignment = la

    with pd.ExcelWriter(path_xl, engine='openpyxl') as writer:
        wb = writer.book

        # ── H1: Resumen Ejecutivo ─────────────────────────────
        ws1 = wb.create_sheet('Resumen')
        ws1['A1'] = 'MVO PORTFOLIO ANALYSIS  —  RESUMEN EJECUTIVO'
        ws1['A1'].font = Font(name='Calibri', bold=True, size=14,
                               color='E6EDF3')
        ws1['A1'].fill = PatternFill('solid', fgColor='0D1117')
        ws1.merge_cells('A1:L1')

        ws1['A3'] = f'Período de análisis:'
        ws1['B3'] = f'{ret.index[0].date()} → {ret.index[-1].date()}'
        ws1['A4'] = 'Observaciones (retornos):'
        ws1['B4'] = ret.shape[0]
        ws1['A5'] = 'Activos:'
        ws1['B5'] = len(tickers)
        ws1['A6'] = 'Portafolios generados:'
        ws1['B6'] = ', '.join(port_names)
        ws1['A7'] = 'RF anual (%):'
        ws1['B7'] = RF_ANUAL * 100
        ws1['A8'] = 'Benchmark:'
        ws1['B8'] = BENCHMARK
        ws1['A9'] = 'Cap máx. por activo (%):'
        ws1['B9'] = CAP_MAX * 100

        ws1['A11'] = 'Portafolio'
        keys_res   = ['ret_anual_%','vol_anual_%','sharpe','sortino',
                       'calmar','mdd_%','var_hist_95%','cvar_hist_95%',
                       'beta','jensen_alpha_%','diversif_ratio','hhi_%']
        labels_res = ['Ret Anual %','Vol Anual %','Sharpe','Sortino',
                       'Calmar','Max DD %','VaR 95%','CVaR 95%',
                       'Beta','Jensen α %','Div. Ratio','HHI %']

        for c_i, lbl in enumerate(labels_res, start=2):
            cell = ws1.cell(row=11, column=c_i, value=lbl)
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = ca

        ws1.cell(row=11, column=1, value='Portafolio').fill = hdr_fill
        ws1.cell(row=11, column=1).font = hdr_font
        ws1.cell(row=11, column=1).alignment = ca

        for r_i, pn in enumerate(port_names, start=12):
            ws1.cell(row=r_i, column=1, value=pn).font = Font(
                name='Calibri', bold=True, size=9)
            for c_i, k in enumerate(keys_res, start=2):
                val  = metricas[pn].get(k, np.nan)
                cell = ws1.cell(row=r_i, column=c_i,
                                 value=round(float(val), 4)
                                       if not np.isnan(val) else 'N/A')
                cell.font   = body_font
                cell.border = bord
                cell.alignment = ra
                if isinstance(val, float) and not np.isnan(val):
                    cell.number_format = '0.0000'
        _aw(ws1)

        # ── H2: Precios ────────────────────────────────────────
        precios_xl = precios.copy()
        precios_xl.index = precios_xl.index.strftime('%Y-%m-%d')
        precios_xl.index.name = 'Fecha'
        precios_xl.to_excel(writer, sheet_name='Precios')
        ws_p = writer.sheets['Precios']
        _fmt_hdr(ws_p)
        for row in ws_p.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.0000'
        _aw(ws_p)

        # ── H3: Retornos_Log ───────────────────────────────────
        ret_xl = ret.copy()
        ret_xl.index = ret_xl.index.strftime('%Y-%m-%d')
        ret_xl.index.name = 'Fecha'
        ret_xl.to_excel(writer, sheet_name='Retornos_Log')
        ws_r = writer.sheets['Retornos_Log']
        _fmt_hdr(ws_r)
        for row in ws_r.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '0.000000'
        _aw(ws_r)

        # ── H4: Descriptivos ───────────────────────────────────
        desc_xl = desc.reset_index()
        desc_xl.to_excel(writer, sheet_name='Descriptivos', index=False)
        ws_d = writer.sheets['Descriptivos']
        _fmt_hdr(ws_d)
        hdrs = [c.value for c in ws_d[1]]
        for row in ws_d.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font; cell.border = bord
                hn = hdrs[cell.column - 1] if cell.column <= len(hdrs) else ''
                if 'Rechaza' in str(hn) or 'ARCH' in str(hn):
                    if cell.value is True:
                        cell.fill = neg_fill; cell.font = neg_font
                    elif cell.value is False:
                        cell.fill = pos_fill; cell.font = pos_font
                elif isinstance(cell.value, float):
                    cell.number_format = '0.0000'; cell.alignment = ra
        _aw(ws_d)

        # ── H5: Correlacion ────────────────────────────────────
        corr_xl = corr_sp.reset_index()
        corr_xl.to_excel(writer, sheet_name='Correlacion', index=False)
        ws_c = writer.sheets['Correlacion']
        _fmt_hdr(ws_c)
        n_t = len(tickers)
        for row in ws_c.iter_rows(min_row=2, max_row=n_t+1,
                                   min_col=2, max_col=n_t+1):
            for cell in row:
                cell.font = body_font; cell.border = bord
                if isinstance(cell.value, float):
                    cell.number_format = '0.0000'; cell.alignment = ca
                    v = abs(cell.value)
                    if v >= 0.7:
                        cell.fill = PatternFill('solid', fgColor='FECACA')
                    elif v >= 0.4:
                        cell.fill = PatternFill('solid', fgColor='FEF3C7')
        _aw(ws_c)

        # ── H6: Frontera Eficiente ─────────────────────────────
        if len(fe_vols) > 0:
            fe_df = pd.DataFrame({'Vol_Anual_%': fe_vols,
                                   'Ret_Anual_%': fe_rets})
            fe_df.to_excel(writer, sheet_name='Frontera', index=False)
            ws_f = writer.sheets['Frontera']
            _fmt_hdr(ws_f)
            for row in ws_f.iter_rows(min_row=2):
                for cell in row:
                    cell.font = body_font; cell.border = bord
                    if isinstance(cell.value, float):
                        cell.number_format = '0.0000'; cell.alignment = ra
            _aw(ws_f)

        # ── H7: Pesos ─────────────────────────────────────────
        w_df = pd.DataFrame({pn: weights[pn] for pn in port_names},
                              index=tickers).T
        w_df.index.name = 'Portafolio'
        w_df.to_excel(writer, sheet_name='Pesos')
        ws_w = writer.sheets['Pesos']
        _fmt_hdr(ws_w)
        for row in ws_w.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font; cell.border = bord
                if isinstance(cell.value, float):
                    cell.number_format = '0.00%'
                    cell.alignment = ra
                    if cell.value > 0:
                        cell.fill = PatternFill(
                            'solid', fgColor='DBEAFE')
        _aw(ws_w)

        # ── H8: Métricas completas ────────────────────────────
        all_keys = sorted({k for m in metricas.values() for k in m})
        rows_m   = []
        for pn in port_names:
            row_m = {'Portafolio': pn}
            for k in all_keys:
                row_m[k] = metricas[pn].get(k, np.nan)
            rows_m.append(row_m)
        met_df = pd.DataFrame(rows_m).set_index('Portafolio')
        met_df.to_excel(writer, sheet_name='Metricas')
        ws_m = writer.sheets['Metricas']
        _fmt_hdr(ws_m)
        for row in ws_m.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font; cell.border = bord
                if isinstance(cell.value, float):
                    cell.number_format = '0.0000'; cell.alignment = ra
        _aw(ws_m)

        # ── H9: VaR / CVaR ────────────────────────────────────
        var_rows = []
        for pn in port_names:
            row_v = {'Portafolio': pn}
            for a in ALPHAS:
                lbl = int(a*100)
                for k in [f'var_hist_{lbl}%', f'var_norm_{lbl}%',
                           f'cvar_hist_{lbl}%']:
                    row_v[k] = metricas[pn].get(k, np.nan)
            var_rows.append(row_v)
        var_df = pd.DataFrame(var_rows).set_index('Portafolio')
        var_df.to_excel(writer, sheet_name='VaR_CVaR')
        ws_v = writer.sheets['VaR_CVaR']
        _fmt_hdr(ws_v)
        for row in ws_v.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font; cell.border = bord
                if isinstance(cell.value, float):
                    cell.number_format = '0.0000'; cell.alignment = ra
        _aw(ws_v)

        # ── H10: MCTR / PCTR ──────────────────────────────────
        mctr_rows = []
        for pn in port_names:
            for tk in tickers:
                mctr_rows.append({
                    'Portafolio': pn,
                    'Activo':     tk,
                    'Peso_%':     weights[pn][tickers.index(tk)] * 100,
                    'MCTR':       mctr_d[pn].get(tk, 0),
                    'PCTR_%':     pctr_d[pn].get(tk, 0) * 100,
                })
        mctr_df = pd.DataFrame(mctr_rows)
        mctr_df.to_excel(writer, sheet_name='MCTR_PCTR', index=False)
        ws_mc = writer.sheets['MCTR_PCTR']
        _fmt_hdr(ws_mc)
        for row in ws_mc.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font; cell.border = bord
                if isinstance(cell.value, float):
                    cell.number_format = '0.0000'; cell.alignment = ra
        _aw(ws_mc)

    print(f"\n  ✓ Excel generado: {os.path.basename(path_xl)}")
    return path_xl

# ──────────────────────────────────────────────────────────────
# 13. MAIN
# ──────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  MVO PORTFOLIO CONSTRUCTION & RISK ANALYSIS  v2.0")
    print("=" * 60)

    # ── Modo de entrada ───────────────────────────────────────
    modo = gui_modo()
    print(f"\n  Modo seleccionado: {'A (Excel/CSV)' if modo=='A' else 'B (yfinance)'}")

    # ── Carpeta de salida ─────────────────────────────────────
    out_dir = gui_carpeta("Selecciona la carpeta de salida")
    ts      = datetime.now().strftime('%Y%m%d_%H%M')
    base    = f'mvo_{ts}'
    print(f"  Carpeta salida: {out_dir}")

    # ── Adquisición de datos ──────────────────────────────────
    print("\n[1/7] ADQUISICIÓN DE DATOS")
    print("-" * 50)

    if modo == 'A':
        ruta    = gui_archivo("Selecciona archivo de precios (Excel/CSV)")
        print(f"  Archivo: {os.path.basename(ruta)}")
        precios = cargar_excel_csv(ruta)

    else:  # modo B — yfinance
        if not YF_OK:
            print("  !! yfinance no instalado. pip install yfinance")
            sys.exit(1)
        tks_str = gui_string(
            "Tickers", "Ingresa los tickers separados por comas:",
            "AAPL,MSFT,GOOGL,NVDA,LLY,JPM,XOM,WMT")
        tickers_in = [t.strip().upper() for t in tks_str.split(',') if t.strip()]
        hoy     = datetime.today()
        d_ini   = gui_string("Fecha inicio", "Fecha inicio (YYYY-MM-DD):",
                             (hoy - timedelta(days=5*365)).strftime('%Y-%m-%d'))
        d_fin   = gui_string("Fecha fin", "Fecha fin (YYYY-MM-DD):",
                              hoy.strftime('%Y-%m-%d'))
        print(f"  Tickers: {tickers_in}")
        print(f"  Período: {d_ini} → {d_fin}")
        precios = cargar_yfinance(tickers_in, d_ini, d_fin)

    # ── QC + retornos ─────────────────────────────────────────
    print("\n[2/7] CONTROL DE CALIDAD + RETORNOS")
    print("-" * 50)
    precios, ret = qc_y_retornos(precios)
    tickers      = list(ret.columns)
    n_act        = len(tickers)

    # ── Benchmark ─────────────────────────────────────────────
    bk_ticker = gui_string("Benchmark", "Ticker del benchmark:", BENCHMARK)
    fecha_ini = str(ret.index[0].date())
    fecha_fin = str((ret.index[-1] + pd.Timedelta(days=1)).date())
    mkt_ret   = descargar_benchmark(fecha_ini, fecha_fin, bk_ticker)
    if mkt_ret is not None:
        print(f"  ✓ Benchmark {bk_ticker}: {len(mkt_ret)} obs descargadas")
    else:
        print(f"  ⚠ Benchmark no disponible — métricas CAPM omitidas")

    # RF global
    global RF_ANUAL
    RF_ANUAL = gui_float("Tasa libre de riesgo",
                          "Tasa libre de riesgo anual (ej. 0.0525 = 5.25%):",
                          RF_ANUAL)
    rf_d = (1 + RF_ANUAL) ** (1 / ANNUAL) - 1
    print(f"  RF anual: {RF_ANUAL*100:.2f}%  (diaria: {rf_d*100:.4f}%)")

    # ── Descriptivos ──────────────────────────────────────────
    print("\n[3/7] ANÁLISIS DESCRIPTIVO Y DISTRIBUCIONAL")
    print("-" * 50)
    desc = analisis_descriptivo(ret)

    # ── Correlación ───────────────────────────────────────────
    print("\n[4/7] CORRELACIÓN DE SPEARMAN")
    print("-" * 50)
    corr_sp, pval_sp, corr_ord, pval_ord, linkage, ord_idx = \
        analisis_correlacion(ret)

    # ── Parámetros MVO ────────────────────────────────────────
    mu    = ret.mean().values             # retornos medios diarios
    Sigma = ret.cov().values              # covarianza diaria
    R_mat = ret.values                    # matriz T×N (retornos históricos)

    # ── Optimización ──────────────────────────────────────────
    print("\n[5/7] OPTIMIZACIÓN MVO")
    print("-" * 50)
    weights = optimizar_portafolios(mu, Sigma, rf_d, R_mat, tickers)
    port_names = list(weights.keys())

    print("\n  Calculando frontera eficiente...")
    fe_vols, fe_rets, fe_pesos = frontera_eficiente(mu, Sigma, rf_d,
                                                      N_FRONTIER)
    print(f"  ✓ {len(fe_vols)} puntos en la frontera")

    # Retornos diarios de cada portafolio
    port_returns = pd.DataFrame(index=ret.index)
    for pn in port_names:
        port_returns[pn] = ret.values @ weights[pn]

    # ── Métricas de riesgo ────────────────────────────────────
    print("\n[6/7] MÉTRICAS DE RIESGO")
    print("-" * 50)
    metricas  = {}
    mctr_dict = {}
    pctr_dict = {}

    for pn in port_names:
        w = weights[pn]
        mctr_s, pctr_s = calcular_mctr(w, Sigma, tickers)

        m = calcular_metricas(port_returns[pn], mkt_ret, rf_d)

        # Diversification Ratio = Σ(w_i × σ_i) / σ_p
        sigma_ind = ret.std().values
        sigma_p   = _sigma_p(w, Sigma)
        dr        = (w @ sigma_ind) / sigma_p if sigma_p > 0 else np.nan
        m['diversif_ratio'] = dr

        # HHI de pesos y número efectivo de activos
        m['hhi_%']    = float(np.sum(w**2) * 100)
        m['n_ef']     = float(1 / np.sum(w**2)) if np.sum(w**2) > 0 else np.nan

        metricas[pn]  = m
        mctr_dict[pn] = mctr_s
        pctr_dict[pn] = pctr_s

        s_val = m.get('sharpe', np.nan)
        v_val = m.get('vol_anual_%', np.nan)
        r_val = m.get('ret_anual_%', np.nan)
        print(f"    {pn:<6}  Ret={r_val:.1f}%  Vol={v_val:.1f}%  "
              f"Sharpe={s_val:.3f}  DR={dr:.3f}")

    # ── Visualizaciones ───────────────────────────────────────
    print("\n[7/7] VISUALIZACIONES + EXCEL")
    print("-" * 50)
    imgs = []
    imgs.append(grafico_distribuciones(ret, desc, out_dir, base))
    imgs.append(grafico_correlacion(corr_ord, pval_ord, out_dir, base))
    imgs.append(grafico_frontera(fe_vols, fe_rets, metricas, out_dir, base))
    imgs.append(grafico_pesos(weights, tickers, out_dir, base))
    imgs.append(grafico_drawdown(port_returns, port_names, out_dir, base))
    imgs.append(grafico_metricas_resumen(metricas, out_dir, base))

    # MCTR (solo si n_act ≤ 20 para legibilidad)
    if n_act <= 20:
        # Preparar datos de pesos como Series para la función
        mctr_plot = {pn: (mctr_dict[pn], pctr_dict[pn]) for pn in port_names}
        grafico_mctr(
            {pn: pd.Series(weights[pn], index=tickers) for pn in port_names},
            pctr_dict, tickers, out_dir, base)

    # Excel
    exportar_excel(
        precios, ret, desc, corr_sp,
        fe_vols, fe_rets, weights, metricas,
        mctr_dict, pctr_dict,
        tickers, port_names, out_dir, base, imgs
    )

    # ── Resumen final ─────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  ANÁLISIS COMPLETADO")
    print("=" * 60)
    print(f"  Activos analizados   : {n_act}")
    print(f"  Portafolios generados: {', '.join(port_names)}")
    print(f"  Puntos frontera      : {len(fe_vols)}")
    print(f"\n  {'Archivo':<50} {'KB':>7}")
    print("  " + "-"*58)
    for fname in os.listdir(out_dir):
        if fname.startswith(base):
            fp = os.path.join(out_dir, fname)
            print(f"  {fname:<50} {os.path.getsize(fp)/1024:>7.1f}")
    print(f"\n  Carpeta: {out_dir}")
    print("=" * 60)


# ──────────────────────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────────────────────
if __name__ == '__main__':
    main()