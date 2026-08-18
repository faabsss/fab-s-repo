# =============================================================================
# ANÁLISIS DE PRECIOS BURSÁTILES
# Extracción · QC · Retornos Log · Estacionariedad · ARCH-LM · GARCH(1,1)
# =============================================================================
# Instalaciones (descomentar en Google Colab)
# !pip install --upgrade yfinance pandas numpy matplotlib seaborn scipy statsmodels openpyxl arch

import os
import sys
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker as mticker
import seaborn as sns
from scipy import stats
import yfinance as yf
from datetime import datetime, timedelta
from statsmodels.tsa.stattools import adfuller, kpss
from statsmodels.stats.diagnostic import het_arch

try:
    from arch import arch_model
    from arch.unitroot import PhillipsPerron
    ARCH_OK = True
except ImportError:
    ARCH_OK = False

warnings.filterwarnings("ignore")
IN_COLAB = "google.colab" in sys.modules

# =============================================================================
# TEMA CLARO — CONSTANTES Y PALETA
# =============================================================================
LBG    = "#F8F9FA"   # fondo figura (gris muy suave)
LAX    = "#FFFFFF"   # fondo ejes
LTXT   = "#212529"   # texto principal
LMUT   = "#6C757D"   # texto secundario / ticks
LGRD   = "#E9ECEF"   # cuadrícula
LSPIN  = "#CED4DA"   # bordes de ejes

# Paleta académica: Seaborn Deep extendida — sofisticada, no saturada
_COLORES_BASE = [
    "#4C72B0",  # azul pizarra
    "#C44E52",  # rojo polvo
    "#55A868",  # verde salvia
    "#8172B3",  # violeta suave
    "#CCB974",  # arena dorada
    "#64B5CD",  # azul cielo
    "#DD8452",  # terracota
    "#937860",  # marrón cálido
    "#2E7D32",  # verde bosque
    "#1565C0",  # azul real
    "#AD1457",  # rosa profundo
    "#E65100",  # naranja tostado
    "#546E7A",  # pizarra
    "#F9A825",  # ámbar
    "#6A1B9A",  # púrpura oscuro
]

plt.rcParams.update({
    "figure.dpi"         : 150,
    "savefig.dpi"        : 300,
    "font.family"        : "sans-serif",
    # Para usar Poppins: instala la fuente en el sistema y ponla de primero
    "font.sans-serif"    : ["Arial", "DejaVu Sans", "Helvetica", "sans-serif"],
    "axes.titlesize"     : 11,
    "axes.titleweight"   : "bold",
    "axes.labelsize"     : 9,
    "xtick.labelsize"    : 8,
    "ytick.labelsize"    : 8,
    "legend.fontsize"    : 8,
    "figure.facecolor"   : LBG,
    "axes.facecolor"     : LAX,
    "savefig.facecolor"  : LBG,
    "text.color"         : LTXT,
    "axes.labelcolor"    : LTXT,
    "xtick.color"        : LMUT,
    "ytick.color"        : LMUT,
    "axes.edgecolor"     : LSPIN,
    "legend.facecolor"   : LBG,
    "legend.edgecolor"   : LSPIN,
    "legend.framealpha"  : 0.9,
    "axes.grid"          : True,
    "grid.color"         : LGRD,
    "grid.linestyle"     : "--",
    "grid.linewidth"     : 0.55,
    "grid.alpha"         : 0.8,
    "axes.spines.top"    : False,
    "axes.spines.right"  : False,
})


def construir_paleta(tickers: list) -> dict:
    return {t: _COLORES_BASE[i % len(_COLORES_BASE)] for i, t in enumerate(tickers)}


def _guardar(fig, output_dir: str, base_name: str, sufijo: str) -> None:
    path = os.path.join(output_dir, f"{base_name}_{sufijo}.png")
    fig.savefig(path, dpi=300, bbox_inches="tight",
                facecolor=LBG, edgecolor="none")
    print(f"    OK  {os.path.basename(path)}")
    plt.show()
    plt.close(fig)


# =============================================================================
# INPUT DEL USUARIO
# =============================================================================
def solicitar_configuracion() -> tuple:
    print("\n" + "=" * 60)
    print("   ANALISIS DE PORTAFOLIO BURSATIL — SERIE DIARIA")
    print("=" * 60)
    print("\nIngresa las siglas bursátiles separadas por comas")
    entrada = input("  Tickers: ")
    tickers = [t.strip().upper() for t in entrada.split(",") if t.strip()]
    if not tickers:
        raise ValueError("No se ingresó ningún ticker.")

    print("\nHorizonte temporal")
    print("[Enter] para usar los últimos 5 años calendario")
    fecha_ini = input("  Fecha inicio (YYYY-MM-DD) o Enter: ").strip()
    fecha_fin = input("  Fecha fin    (YYYY-MM-DD) o Enter para hoy: ").strip()

    end   = fecha_fin if fecha_fin else datetime.today().strftime("%Y-%m-%d")
    start = fecha_ini if fecha_ini else (
        datetime.today() - timedelta(days=5 * 365)
    ).strftime("%Y-%m-%d")

    print("\n  Validando tickers en yfinance...")
    nombres, validos = {}, []
    for t in tickers:
        try:
            info   = yf.Ticker(t).info
            nombre = info.get("longName") or info.get("shortName") or t
            nombres[t] = nombre
            validos.append(t)
            print(f"    OK  {t:8s} -> {nombre}")
        except Exception:
            print(f"    !!  {t} no encontrado. Se omitirá.")

    if not validos:
        raise ValueError("Ningún ticker válido encontrado.")
    print(f"\n  Periodo : {start} -> {end}  |  Activos : {', '.join(validos)}")
    return validos, start, end, nombres


def obtener_destino() -> tuple:
    if IN_COLAB:
        print("\n[Colab] Ingresa la ruta manualmente.")
        output_dir = input("  Carpeta destino: ").strip() or "/content"
        base_name  = input("  Nombre base (sin extension): ").strip() or "portafolio"
    else:
        try:
            import tkinter as tk
            from tkinter import filedialog, simpledialog
            root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
            output_dir = filedialog.askdirectory(title="Carpeta de destino") or os.getcwd()
            base_name  = simpledialog.askstring(
                title="Nombre de archivos",
                prompt="Nombre base (sin extension):",
                initialvalue="portafolio_analisis",
            ) or "portafolio_analisis"
            root.destroy()
        except Exception as exc:
            print(f"  GUI no disponible ({exc}).")
            output_dir = input("  Carpeta destino: ").strip() or os.getcwd()
            base_name  = input("  Nombre base: ").strip() or "portafolio_analisis"

    os.makedirs(output_dir, exist_ok=True)
    print(f"  Destino : {output_dir}  |  Nombre : {base_name}")
    return output_dir, base_name


# =============================================================================
# 1. DESCARGA
# =============================================================================
def descargar_datos(tickers, start, end) -> pd.DataFrame:
    print("\n[1/9] DESCARGA DE DATOS")
    print("-" * 50)
    raw = yf.download(tickers, start=start, end=end, auto_adjust=True, progress=True)
    if isinstance(raw["Close"], pd.Series):
        precios = raw["Close"].to_frame(name=tickers[0])
    else:
        precios = raw["Close"].copy()[tickers]
    precios.index = pd.to_datetime(precios.index)
    precios.columns.name = None
    print(f"  Periodo : {precios.index[0].date()} -> {precios.index[-1].date()}")
    print(f"  Obs.    : {len(precios)} dias x {len(tickers)} activos")
    return precios


# =============================================================================
# 2. CONTROL DE CALIDAD
# =============================================================================
def control_calidad(precios: pd.DataFrame) -> pd.DataFrame:
    print("\n[2/9] CONTROL DE CALIDAD")
    print("-" * 50)
    nulos = precios.isnull().sum()
    print("  Nulos por activo (antes):")
    for t, n in nulos.items():
        print(f"    {t:8s}: {'OK' if n == 0 else f'!! {n} nulos'}")
    precios_c = precios.ffill(limit=1)
    if precios_c.isnull().sum().sum() == 0:
        print("  Sin nulos tras limpieza. OK")
    print("  Rango de precios (min / max):")
    for t in precios_c.columns:
        print(f"    {t:8s}: ${precios_c[t].min():.2f} -> ${precios_c[t].max():.2f}")
    return precios_c


# =============================================================================
# 3. RETORNOS LOGARÍTMICOS
# =============================================================================
def calcular_retornos(precios: pd.DataFrame) -> pd.DataFrame:
    print("\n[3/9] RETORNOS LOGARÍTMICOS  [r_t = ln(P_t / P_{t-1})]")
    print("-" * 50)
    ret = np.log(precios / precios.shift(1)).dropna()
    print(f"  Observaciones: {len(ret)} por activo")
    return ret


# =============================================================================
# 4. TESTS DE ESTACIONARIEDAD
# =============================================================================
def tests_estacionariedad(precios: pd.DataFrame, retornos: pd.DataFrame) -> pd.DataFrame:
    """
    ADF  — H0: raíz unitaria (NO estacionaria). Rechazar => estacionaria.
    KPSS — H0: estacionaria. Rechazar         => NO estacionaria.  [H0 opuesta al ADF]
    PP   — H0: raíz unitaria (NO estacionaria). Rechazar => estacionaria.
    """
    print("\n[4/9] TESTS DE ESTACIONARIEDAD")
    print("-" * 50)
    if not ARCH_OK:
        print("  !! Phillips-Perron omitido (pip install arch)")

    filas = []
    for tipo, df in [("Niveles (precio)", precios), ("Retornos log", retornos)]:
        print(f"\n  -- {tipo} --")
        for t in df.columns:
            serie = df[t].dropna()
            # ADF
            ar = adfuller(serie, autolag="AIC")
            adf_s, adf_p, adf_l = ar[0], ar[1], ar[2]
            adf_c = "Estacionaria" if adf_p < 0.05 else "Raiz unitaria"
            # KPSS
            kr = kpss(serie, regression="c", nlags="auto")
            kp_s, kp_p, kp_l = kr[0], kr[1], kr[2]
            kp_c = "Raiz unitaria" if kp_p < 0.05 else "Estacionaria"
            # PP
            if ARCH_OK:
                pr = PhillipsPerron(serie)
                pp_s, pp_p, pp_l = pr.stat, pr.pvalue, pr.lags
                pp_c = "Estacionaria" if pp_p < 0.05 else "Raiz unitaria"
            else:
                pp_s = pp_p = pp_l = float("nan")
                pp_c = "N/A"
            # Diagnóstico por votos
            votos   = sum([adf_p < 0.05, kp_p >= 0.05, (pp_p < 0.05) if ARCH_OK else False])
            n_tests = 3 if ARCH_OK else 2
            diag = ("Estacionaria (unanime)"   if votos == n_tests else
                    "No estacionaria (unanime)" if votos == 0       else
                    "Resultado mixto")
            print(f"    {t:6s} | ADF p={adf_p:.4f} [{adf_c:13s}]  "
                  f"KPSS p={kp_p:.4f} [{kp_c:13s}]  "
                  + (f"PP p={pp_p:.4f} [{pp_c:13s}]" if ARCH_OK else "PP: N/A")
                  + f"  => {diag}")
            filas.append({
                "Ticker": t, "Serie": tipo,
                "ADF Estadístico": round(adf_s, 4), "ADF p-valor": round(adf_p, 4),
                "ADF Lags": int(adf_l), "ADF Conclusión": adf_c,
                "KPSS Estadístico": round(kp_s, 4), "KPSS p-valor": round(kp_p, 4),
                "KPSS Lags": int(kp_l), "KPSS Conclusión": kp_c,
                "PP Estadístico": round(pp_s, 4) if ARCH_OK else float("nan"),
                "PP p-valor": round(pp_p, 4) if ARCH_OK else float("nan"),
                "PP Lags": int(pp_l) if ARCH_OK else float("nan"),
                "PP Conclusión": pp_c,
                "Diagnóstico": diag,
            })
    return pd.DataFrame(filas)


# =============================================================================
# 5. TEST ARCH-LM (efecto ARCH — justificación del GARCH)
# =============================================================================
def test_arch_lm(retornos: pd.DataFrame, lags_list: list = [5, 10, 20]) -> pd.DataFrame:
    """
    Test de multiplicador de Lagrange de Engle (1982).

    H0: NO hay efectos ARCH — la varianza condicional es constante
        (los cuadrados de los residuos no están autocorrelacionados)
    H1: HAY efectos ARCH — la varianza condicional es dinámica
        (justifica el uso de modelos GARCH)

    Rechazar H0 (p < 0.05) => la serie exhibe heteroscedasticidad
    condicional => el modelo GARCH es apropiado.

    Se aplica solo a retornos (ya que la prueba asume media estacionaria).
    """
    print("\n[5/9] TEST ARCH-LM (efecto ARCH)")
    print("-" * 50)
    print("  H0: No hay efectos ARCH (varianza condicional constante)")
    print("  H1: Hay efectos ARCH  => GARCH justificado\n")

    filas = []
    for t in retornos.columns:
        r = retornos[t].dropna().values
        for lag in lags_list:
            lm_s, lm_p, f_s, f_p = het_arch(r, nlags=lag)
            concl = "ARCH presente [GARCH justificado]" if lm_p < 0.05 else "Sin efecto ARCH"
            print(f"    {t:6s} Lag={lag:2d} | LM stat={lm_s:9.3f}  p={lm_p:.4f}  => {concl}")
            filas.append({
                "Ticker"          : t,
                "Lags"            : lag,
                "LM Estadístico"  : round(lm_s, 4),
                "LM p-valor"      : round(lm_p, 4),
                "F Estadístico"   : round(f_s,  4),
                "F p-valor"       : round(f_p,  4),
                "Conclusión"      : concl,
            })
    return pd.DataFrame(filas)


# =============================================================================
# 6. ESTADÍSTICOS DESCRIPTIVOS
# =============================================================================
def estadisticos_descriptivos(retornos: pd.DataFrame, nombres: dict) -> pd.DataFrame:
    print("\n[6/9] ESTADÍSTICOS DESCRIPTIVOS")
    print("-" * 50)
    filas = []
    for t in retornos.columns:
        r = retornos[t].dropna()
        jb_s, jb_p = stats.jarque_bera(r)
        filas.append({
            "Ticker": t, "Nombre": nombres.get(t, t),
            "N obs.": len(r),
            "Media diaria": r.mean(), "Desv. Std diaria": r.std(),
            "Volat. anual": r.std() * np.sqrt(252),
            "Mínimo": r.min(), "Maximo": r.max(),
            "Asimetría": r.skew(), "Curtosis (exceso)": r.kurtosis(),
            "JB Estadístico": jb_s, "JB p-valor": jb_p,
            "Normalidad (5%)": "RECHAZA H0" if jb_p < 0.05 else "No rechaza",
        })
    tabla = pd.DataFrame(filas).set_index("Ticker")
    cols  = ["Media diaria", "Desv. Std diaria", "Volat. anual",
             "Asimetría", "Curtosis (exceso)", "JB p-valor", "Normalidad (5%)"]
    print(f"\n{tabla[cols].round(4).to_string()}")
    return tabla


# =============================================================================
# 7. MODELO GARCH(1,1) — VARIANZA CONDICIONAL + PROYECCIÓN 63 DÍAS
# =============================================================================
def estimar_garch(retornos: pd.DataFrame) -> dict:
    """
    Especificación: GARCH(1,1) con distribución t-Student

    Ecuación de retornos:
        r_t = mu + epsilon_t,   epsilon_t = sigma_t * z_t,   z_t ~ t(nu)

    Ecuación de varianza condicional:
        sigma^2_t = omega + alpha1 * epsilon^2_{t-1} + beta1 * sigma^2_{t-1}

    donde:
        omega  > 0   : componente de varianza de largo plazo
        alpha1 >= 0  : efecto ARCH (impacto del shock de ayer en la varianza de hoy)
        beta1  >= 0  : efecto GARCH (persistencia de la varianza de ayer)
        alpha1 + beta1 < 1 : condición de estacionariedad en covarianza

    Varianza incondicional (largo plazo):
        sigma^2_inf = omega / (1 - alpha1 - beta1)

    Proyeccion h pasos adelante:
        E[sigma^2_{T+h} | I_T] = sigma^2_inf
                                 + (alpha1+beta1)^(h-1) * (sigma^2_{T+1} - sigma^2_inf)

    Semivida del shock de volatilidad:
        t_1/2 = log(0.5) / log(alpha1 + beta1)   [dias]

    Los retornos se escalan x100 (porcentaje) para estabilidad numerica.
    La volatilidad reportada se re-escala y se anualiza: sigma_t * sqrt(252) / 100
    """
    if not ARCH_OK:
        print("\n  !! arch no instalado — GARCH omitido (pip install arch)")
        return {}

    print("\n[7/9] MODELO GARCH(1,1) — distribucion t-Student")
    print("-" * 50)
    print("  Escalando retornos x100 para estabilidad numerica...")

    resultados = {}
    for t in retornos.columns:
        r_pct = retornos[t].dropna() * 100          # porcentaje

        am  = arch_model(r_pct, mean="Constant", vol="Garch", p=1, q=1, dist="t")
        res = am.fit(disp="off", options={"maxiter": 1000})

        mu     = res.params["mu"]
        omega  = res.params["omega"]
        alpha1 = res.params["alpha[1]"]
        beta1  = res.params["beta[1]"]
        nu     = res.params["nu"]
        pers   = alpha1 + beta1

        lr_var       = omega / (1 - pers) if pers < 1 else np.inf
        lr_vol_ann   = (np.sqrt(lr_var) / 100 * np.sqrt(252)) if np.isfinite(lr_var) else np.inf
        half_life    = (np.log(0.5) / np.log(pers)) if 0 < pers < 1 else np.inf

        # Serie de volatilidad condicional (diaria, %)
        cond_vol_pct = res.conditional_volatility            # sigma_t en %
        cond_vol_ann = cond_vol_pct / 100 * np.sqrt(252)     # anualizada

        # Proyección 63 días (= 3 meses bursátiles ≈ 63 ruedas)
        fc      = res.forecast(horizon=63, reindex=False)
        var_fc  = fc.variance.values[-1]                     # sigma^2_{T+h} en %^2
        vol_fc_pct = np.sqrt(var_fc)                         # sigma_{T+h} diaria en %
        vol_fc_ann = vol_fc_pct / 100 * np.sqrt(252)         # anualizada

        resultados[t] = dict(
            res=res, mu=mu, omega=omega, alpha1=alpha1, beta1=beta1,
            nu=nu, pers=pers, lr_var=lr_var, lr_vol_ann=lr_vol_ann,
            half_life=half_life, loglik=res.loglikelihood,
            aic=res.aic, bic=res.bic,
            cond_vol_pct=cond_vol_pct,    # pd.Series, index = retornos.index
            cond_vol_ann=cond_vol_ann,
            var_fc=var_fc, vol_fc_pct=vol_fc_pct, vol_fc_ann=vol_fc_ann,
        )

        hl_str = f"{half_life:.1f} dias" if np.isfinite(half_life) else "inf"
        print(f"    {t:6s}: mu={mu:.6f}%  omega={omega:.4f}  alpha1={alpha1:.4f}  beta1={beta1:.4f}  "
              f"nu={nu:.2f}  persist={pers:.4f}  semivida={hl_str}  AIC={res.aic:.1f}")

    return resultados


# =============================================================================
# 8. GRÁFICOS — TEMA CLARO SOFISTICADO
# =============================================================================

# ── 8.1  Serie en niveles — un gráfico por activo ────────────────────────────
def graficos_serie_niveles(precios, palette, nombres, output_dir, base_name):
    for t in precios.columns:
        s = precios[t].dropna()
        c = palette.get(t, _COLORES_BASE[0])
        fig, ax = plt.subplots(figsize=(11, 4), facecolor=LBG)
        ax.set_facecolor(LAX)
        ax.plot(s.index, s.values, color=c, linewidth=1.5, alpha=0.9)
        ax.fill_between(s.index, s.values, s.min(), color=c, alpha=0.07)
        # Máximo y mínimo
        im, ix = s.idxmax(), s.idxmin()
        ax.scatter([im], [s.max()], color=c, s=50, zorder=5)
        ax.scatter([ix], [s.min()], color="#C44E52", s=50, zorder=5)
        ax.annotate(f"Max ${s.max():.2f}  {im.strftime('%b %Y')}",
                    xy=(im, s.max()), xytext=(10, -15), textcoords="offset points",
                    color=c, fontsize=7.5, fontweight="bold",
                    arrowprops=dict(arrowstyle="-", color=c, lw=0.7))
        ax.annotate(f"Min ${s.min():.2f}  {ix.strftime('%b %Y')}",
                    xy=(ix, s.min()), xytext=(10, 8), textcoords="offset points",
                    color="#C44E52", fontsize=7.5, fontweight="bold",
                    arrowprops=dict(arrowstyle="-", color="#C44E52", lw=0.7))
        ax.set_title(f"{t}  |  {nombres.get(t, t)}", pad=10)
        ax.set_ylabel("Precio ajustado (USD)")
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
        plt.xticks(rotation=30, ha="right")
        ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("$%.0f"))
        fig.suptitle("Serie en niveles — Precio de cierre ajustado diario",
                     color=LMUT, fontsize=8.5, y=1.00)
        plt.tight_layout()
        _guardar(fig, output_dir, base_name, f"precio_{t}")


# ── 8.2  Retornos logarítmicos — un gráfico por activo ───────────────────────
def graficos_serie_retornos(retornos, palette, nombres, output_dir, base_name):
    for t in retornos.columns:
        r  = retornos[t].dropna()
        c  = palette.get(t, _COLORES_BASE[0])
        q05, q95 = r.quantile(0.05), r.quantile(0.95)
        fig, ax = plt.subplots(figsize=(11, 4), facecolor=LBG)
        ax.set_facecolor(LAX)
        ax.bar(r.index, r.values, width=1, color=c, alpha=0.40, linewidth=0)
        ax.fill_between(r.index, r.values, 0, where=(r.values < q05),
                        color="#C44E52", alpha=0.55, label=f"Cola inf. P5  (<{q05:.3f})")
        ax.fill_between(r.index, r.values, 0, where=(r.values > q95),
                        color="#55A868", alpha=0.55, label=f"Cola sup. P95 (>{q95:.3f})")
        ax.axhline(0, color=LMUT, linewidth=0.7, linestyle="--")
        ax.text(0.01, 0.97,
                f"sigma_anual={r.std()*np.sqrt(252):.1%}   "
                f"gamma1={r.skew():.3f}   kappa={r.kurtosis():.2f}",
                transform=ax.transAxes, ha="left", va="top",
                color=LMUT, fontsize=7.5)
        ax.set_title(f"{t}  |  {nombres.get(t, t)}", pad=10)
        ax.set_ylabel("Retorno logaritmico diario")
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
        plt.xticks(rotation=30, ha="right")
        ax.legend(loc="lower right", fontsize=7)
        fig.suptitle("Retornos logaritmicos diarios  [r_t = ln(P_t / P_{t-1})]",
                     color=LMUT, fontsize=8.5, y=1.00)
        plt.tight_layout()
        _guardar(fig, output_dir, base_name, f"retorno_{t}")


# ── 8.3  Distribuciones con KDE ──────────────────────────────────────────────
def grafico_distribuciones(retornos, palette, nombres, output_dir, base_name):
    tickers = list(retornos.columns)
    n = len(tickers); ncols = min(3, n); nrows = (n + ncols - 1) // ncols
    fig, axes = plt.subplots(nrows, ncols, figsize=(13, 3.8 * nrows), facecolor=LBG)
    axes_flat = np.array(axes).flatten()
    for i, t in enumerate(tickers):
        ax = axes_flat[i]; ax.set_facecolor(LAX)
        r  = retornos[t].dropna()
        c  = palette.get(t, _COLORES_BASE[0])
        ax.hist(r, bins=80, density=True, color=c, alpha=0.28, edgecolor="none")
        kde_x = np.linspace(r.min(), r.max(), 400)
        ax.plot(kde_x, stats.gaussian_kde(r)(kde_x), color=c, lw=2.0, label="KDE empirica")
        ax.plot(kde_x, stats.norm.pdf(kde_x, r.mean(), r.std()),
                color=LTXT, lw=1.2, ls="--", alpha=0.55, label="Normal teorica")
        ax.text(0.97, 0.95, f"kappa={r.kurtosis():.2f}\ngamma1={r.skew():.2f}",
                transform=ax.transAxes, ha="right", va="top", fontsize=8,
                color=c, fontweight="bold",
                bbox=dict(boxstyle="round,pad=0.3", fc=LBG, ec=c, alpha=0.85))
        ax.set_title(f"{t} — {nombres.get(t, t)}", fontsize=9)
        ax.set_xlabel("Retorno log diario", fontsize=8)
        ax.set_ylabel("Densidad", fontsize=8)
        ax.legend(fontsize=7, frameon=False)
    for j in range(i + 1, len(axes_flat)):
        axes_flat[j].set_visible(False)
    fig.suptitle("Distribución de retornos logarítmicos diarios\n"
                 "(kappa = curtosis exceso   gamma1 = asimetría)", y=1.01)
    plt.tight_layout()
    _guardar(fig, output_dir, base_name, "distribuciones")


# ── 8.4  Heatmap de correlación con texto adaptativo ─────────────────────────
def grafico_heatmap_correlacion(retornos, output_dir, base_name):
    corr = retornos.corr()
    n    = len(corr)
    mask = np.triu(np.ones_like(corr, dtype=bool), k=1)
    cmap = sns.diverging_palette(220, 10, s=85, l=50, as_cmap=True)
    fig_w = max(6.5, n * 1.15); fig_h = max(5.5, n * 1.0)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h), facecolor=LBG)
    ax.set_facecolor(LAX)
    sns.heatmap(corr, mask=mask, cmap=cmap, vmin=-1, vmax=1, center=0,
                annot=False, linewidths=0.8, linecolor=LBG,
                square=True, ax=ax, cbar_kws={"shrink": 0.72, "pad": 0.02})
    # Texto adaptativo por luminancia
    for i in range(n):
        for j in range(n):
            if mask[i, j]:
                continue
            val = corr.iloc[i, j]
            r_bg, g_bg, b_bg, _ = cmap((val + 1) / 2)
            lum  = 0.299 * r_bg + 0.587 * g_bg + 0.114 * b_bg
            ctxt = "#1A1A2E" if lum > 0.45 else "white"
            ax.text(j + 0.5, i + 0.5, f"{val:.2f}",
                    ha="center", va="center",
                    fontsize=10, fontweight="bold", color=ctxt)
    ax.set_title("Matriz de correlacion — Retornos logaritmicos diarios\n"
                 "(triangulo inferior  |  paleta azul-rojo)", pad=12)
    ax.tick_params(colors=LMUT)
    ax.set_xticklabels(ax.get_xticklabels(), rotation=30, ha="right", fontsize=9)
    ax.set_yticklabels(ax.get_yticklabels(), rotation=0, fontsize=9)
    cbar = ax.collections[0].colorbar
    if cbar:
        cbar.ax.tick_params(labelsize=8)
        cbar.set_label("Correlacion de Pearson", fontsize=8)
    plt.tight_layout()
    _guardar(fig, output_dir, base_name, "correlacion")


# ── 8.5  GARCH: retornos + volatilidad condicional — por activo ──────────────
def graficos_garch_volatilidad(retornos, garch_res, palette, nombres,
                                output_dir, base_name):
    if not garch_res:
        return
    for t in retornos.columns:
        if t not in garch_res:
            continue
        r   = retornos[t].dropna()
        c   = palette.get(t, _COLORES_BASE[0])
        cv  = garch_res[t]["cond_vol_ann"]   # vol anualizada (proporción)
        lr  = garch_res[t]["lr_vol_ann"]     # vol incondicional
        pers = garch_res[t]["pers"]

        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(11, 6), sharex=True,
                                        facecolor=LBG,
                                        gridspec_kw={"height_ratios": [1, 1.3],
                                                      "hspace": 0.08})
        for ax in (ax1, ax2):
            ax.set_facecolor(LAX)

        # Panel superior: retornos
        ax1.bar(r.index, r.values, width=1, color=c, alpha=0.45, linewidth=0)
        ax1.axhline(0, color=LMUT, linewidth=0.6, linestyle="--")
        ax1.set_ylabel("Retorno log diario", fontsize=9)
        ax1.set_title(f"{t}  |  {nombres.get(t, t)}", pad=10)

        # Panel inferior: volatilidad condicional anualizada
        ax2.plot(cv.index, cv.values * 100, color=c, linewidth=1.3, alpha=0.9,
                 label=r"$\hat{\sigma}_t$ (volatilidad cond. anualizada)")
        ax2.fill_between(cv.index, cv.values * 100, 0, color=c, alpha=0.08)
        if np.isfinite(lr):
            ax2.axhline(lr * 100, color="#C44E52", linewidth=1.1,
                        linestyle="--", alpha=0.8,
                        label=f"Varianza incond. = {lr*100:.1f}%")
        ax2.set_ylabel("Volatilidad anualizada (%)", fontsize=9)
        ax2.yaxis.set_major_formatter(mticker.PercentFormatter())
        ax2.legend(loc="upper right", fontsize=7.5)
        ax2.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
        ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
        plt.xticks(rotation=30, ha="right")

        fig.suptitle(
            f"GARCH(1,1)-t | alpha1+beta1 (persistencia) = {pers:.4f}",
            fontsize=9, color=LMUT, y=1.00
        )
        plt.tight_layout()
        _guardar(fig, output_dir, base_name, f"garch_volat_{t}")


# ── 8.6  Proyección GARCH: 63 días bursátiles — todos los activos ─────────────
def grafico_proyeccion_garch(garch_res, palette, output_dir, base_name):
    if not garch_res:
        return
    tickers = list(garch_res.keys())
    horizonte = np.arange(1, 64)                # 1 a 63

    fig, ax = plt.subplots(figsize=(11, 5), facecolor=LBG)
    ax.set_facecolor(LAX)

    for t in tickers:
        c   = palette.get(t, _COLORES_BASE[0])
        vol = garch_res[t]["vol_fc_ann"] * 100   # anualizada en %
        lr  = garch_res[t]["lr_vol_ann"] * 100   # incondicional en %

        ax.plot(horizonte, vol, color=c, linewidth=2.0, label=t, alpha=0.9)
        if np.isfinite(lr):
            ax.axhline(lr, color=c, linewidth=0.8, linestyle=":", alpha=0.5)

    ax.set_title(
        "Proyeccion de volatilidad condicional — GARCH(1,1)\n"
        "Horizonte: 63 dias bursatiles (~3 meses)  |  Linea punteada = varianza incondicional"
    )
    ax.set_xlabel("Dias bursatiles hacia adelante")
    ax.set_ylabel("Volatilidad anualizada proyectada (%)")
    ax.yaxis.set_major_formatter(mticker.PercentFormatter())
    ax.set_xlim(1, 63)
    ax.legend(loc="upper right", fontsize=8)
    ax.axvline(21, color=LMUT, linewidth=0.6, linestyle="--", alpha=0.4)
    ax.axvline(42, color=LMUT, linewidth=0.6, linestyle="--", alpha=0.4)
    ax.text(21, ax.get_ylim()[0], "1M", ha="center", va="bottom",
            fontsize=7, color=LMUT)
    ax.text(42, ax.get_ylim()[0], "2M", ha="center", va="bottom",
            fontsize=7, color=LMUT)
    ax.text(63, ax.get_ylim()[0], "3M", ha="center", va="bottom",
            fontsize=7, color=LMUT)

    plt.tight_layout()
    _guardar(fig, output_dir, base_name, "garch_proyeccion")


# =============================================================================
# 9. EXPORTACIÓN A EXCEL  — 7 HOJAS
# =============================================================================
def exportar_excel(precios, retornos, tabla_stats, tabla_tests,
                   tabla_arch, garch_res, output_dir, base_name) -> str:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    path_excel = os.path.join(output_dir, f"{base_name}.xlsx")

    # ── Preparar series ───────────────────────────────────────────────────
    px = precios.copy(); px.index = px.index.strftime("%Y-%m-%d"); px.index.name = "Fecha"
    rx = retornos.copy(); rx.index = rx.index.strftime("%Y-%m-%d"); rx.index.name = "Fecha"

    # ── Hoja GARCH: varianza condicional ──────────────────────────────────
    if garch_res:
        tickers_g = list(garch_res.keys())
        # Alinear todas las series al mismo índice
        idx_g = retornos[tickers_g[0]].dropna().index
        df_cv = pd.DataFrame(index=idx_g)
        for t in tickers_g:
            cv = garch_res[t]["cond_vol_ann"]
            df_cv[f"{t}_vol_cond_anual"] = cv.reindex(idx_g)
        df_cv.index = df_cv.index.strftime("%Y-%m-%d")
        df_cv.index.name = "Fecha"

        # Proyección 63 días
        # Almacenamos como proporción (0.40 = 40%) para que el formato
        # "0.00%" de Excel multiplique por 100 exactamente una sola vez.
        df_proj = pd.DataFrame({"Horizonte (dias)": np.arange(1, 64)})
        for t in tickers_g:
            df_proj[f"{t}_vol_proj_anual"] = garch_res[t]["vol_fc_ann"]

        # Parámetros
        rows_p = []
        for t in tickers_g:
            g = garch_res[t]
            rows_p.append({
                "Ticker"             : t,
                "omega"              : round(g["omega"],    6),
                "alpha1"             : round(g["alpha1"],   6),
                "beta1"              : round(g["beta1"],    6),
                "nu (grados lib.)"   : round(g["nu"],       4),
                "Persistencia"       : round(g["pers"],     6),
                "Semivida (dias)"    : round(g["half_life"],2) if np.isfinite(g["half_life"]) else "inf",
                "Var. incond. (%^2)" : round(g["lr_var"],   6) if np.isfinite(g["lr_var"]) else "inf",
                "Vol. incond. anual" : round(g["lr_vol_ann"]*100, 4) if np.isfinite(g["lr_vol_ann"]) else "inf",
                "Log-Likelihood"     : round(g["loglik"],   4),
                "AIC"                : round(g["aic"],      4),
                "BIC"                : round(g["bic"],      4),
            })
        df_params = pd.DataFrame(rows_p)
    else:
        df_cv     = pd.DataFrame()
        df_proj   = pd.DataFrame()
        df_params = pd.DataFrame()

    # ── Escribir hojas ────────────────────────────────────────────────────
    with pd.ExcelWriter(path_excel, engine="openpyxl") as writer:
        px.to_excel(writer,                      sheet_name="Precios_Ajustados")
        rx.to_excel(writer,                      sheet_name="Retornos_Log")
        tabla_stats.reset_index().to_excel(writer, sheet_name="Estadisticos",   index=False)
        tabla_tests.to_excel(writer,             sheet_name="Tests_Estacion",  index=False)
        tabla_arch.to_excel(writer,              sheet_name="Test_ARCH_LM",    index=False)
        if not df_params.empty:
            df_params.to_excel(writer,           sheet_name="GARCH_Parametros",index=False)
            df_cv.to_excel(writer,               sheet_name="GARCH_Var_Cond")
            df_proj.to_excel(writer,             sheet_name="GARCH_Proyeccion",index=False)

        wb = writer.book

        # ── Estilos ───────────────────────────────────────────────────────
        hf = PatternFill("solid", fgColor="264653")   # verde pizarra oscuro
        hfont = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        bf    = Font(name="Calibri", size=9)
        ca    = Alignment(horizontal="center", vertical="center")
        ra    = Alignment(horizontal="right",  vertical="center")
        thin  = Side(style="thin", color="DEE2E6")
        bord  = Border(bottom=thin)

        def _aw(ws):
            for col in ws.columns:
                ml = max((len(str(c.value)) for c in col if c.value is not None), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 28)

        def _fmt(ws, num_fmt=None):
            for cell in ws[1]:
                cell.fill = hf; cell.font = hfont; cell.alignment = ca
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = bf; cell.border = bord
                    cell.alignment = (ra if (num_fmt and cell.column > 1) else ca)
                    if num_fmt and cell.column > 1:
                        cell.number_format = num_fmt
            _aw(ws); ws.freeze_panes = "B2"

        _fmt(wb["Precios_Ajustados"], "#,##0.0000")
        _fmt(wb["Retornos_Log"],      "0.000000")

        # Estadísticos
        ws_s  = wb["Estadisticos"]
        hdrs  = [c.value for c in ws_s[1]]
        fmts  = {"N obs.": "0", "Media diaria": "0.000000",
                 "Desv. Std diaria": "0.000000", "Volat. anual": "0.00%",
                 "Minimo": "0.000000", "Maximo": "0.000000",
                 "Asimetria": "0.0000", "Curtosis (exceso)": "0.0000",
                 "JB Estadistico": "#,##0.00", "JB p-valor": "0.0000"}
        for cell in ws_s[1]:
            cell.fill = hf; cell.font = hfont; cell.alignment = ca
        for row in ws_s.iter_rows(min_row=2):
            for cell in row:
                cn = hdrs[cell.column - 1]
                cell.font = bf; cell.border = bord
                cell.alignment = ra if cn in fmts else ca
                if cn in fmts: cell.number_format = fmts[cn]
                if cn == "Normalidad (5%)" and cell.value == "RECHAZA H0":
                    cell.fill = PatternFill("solid", fgColor="FFE4E4")
                    cell.font = Font(name="Calibri", bold=True, color="C44E52", size=9)
        _aw(ws_s); ws_s.freeze_panes = "A2"

        # Tests estacionariedad — semáforo
        def _semaforo(ws, col_diag="Diagnostico"):
            hdrs_t = [c.value for c in ws[1]]
            for cell in ws[1]: cell.fill = hf; cell.font = hfont; cell.alignment = ca
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cn = hdrs_t[cell.column - 1]
                    cell.font = bf; cell.border = bord; cell.alignment = ca
                    if cn == col_diag and isinstance(cell.value, str):
                        if "unanime" in cell.value and "No " not in cell.value:
                            cell.fill = PatternFill("solid", fgColor="D4EDDA")
                            cell.font = Font(name="Calibri", bold=True, color="155724", size=9)
                        elif "No estacionaria" in cell.value:
                            cell.fill = PatternFill("solid", fgColor="FFE4E4")
                            cell.font = Font(name="Calibri", bold=True, color="C44E52", size=9)
                        elif "mixto" in cell.value:
                            cell.fill = PatternFill("solid", fgColor="FFF3CD")
                            cell.font = Font(name="Calibri", bold=True, color="856404", size=9)
            _aw(ws); ws.freeze_panes = "C2"

        _semaforo(wb["Tests_Estacion"])

        # ARCH-LM — semáforo por conclusión
        ws_a = wb["Test_ARCH_LM"]
        hdrs_a = [c.value for c in ws_a[1]]
        for cell in ws_a[1]: cell.fill = hf; cell.font = hfont; cell.alignment = ca
        for row in ws_a.iter_rows(min_row=2):
            for cell in row:
                cn = hdrs_a[cell.column - 1]
                cell.font = bf; cell.border = bord; cell.alignment = ca
                if "p-valor" in str(cn) and isinstance(cell.value, float):
                    cell.number_format = "0.0000"
                if cn == "Conclusion" and isinstance(cell.value, str):
                    if "justificado" in cell.value:
                        cell.fill = PatternFill("solid", fgColor="D4EDDA")
                        cell.font = Font(name="Calibri", bold=True, color="155724", size=9)
                    else:
                        cell.fill = PatternFill("solid", fgColor="FFF3CD")
                        cell.font = Font(name="Calibri", bold=True, color="856404", size=9)
        _aw(ws_a); ws_a.freeze_panes = "A2"

        # GARCH parámetros
        if "GARCH_Parametros" in wb.sheetnames:
            ws_gp = wb["GARCH_Parametros"]
            hdrs_gp = [c.value for c in ws_gp[1]]
            num_cols_gp = {"omega", "alpha1", "beta1", "nu (grados lib.)",
                           "Persistencia", "Semivida (dias)",
                           "Var. incond. (%^2)", "Vol. incond. anual",
                           "Log-Likelihood", "AIC", "BIC"}
            for cell in ws_gp[1]: cell.fill = hf; cell.font = hfont; cell.alignment = ca
            for row in ws_gp.iter_rows(min_row=2):
                for cell in row:
                    cn = hdrs_gp[cell.column - 1]
                    cell.font = bf; cell.border = bord
                    cell.alignment = ra if cn in num_cols_gp else ca
                    if cn in num_cols_gp and isinstance(cell.value, float):
                        cell.number_format = "0.000000"
                    # Resaltar si persistencia >= 0.98 (cuasi-integrado)
                    if cn == "Persistencia" and isinstance(cell.value, float) and cell.value >= 0.98:
                        cell.fill = PatternFill("solid", fgColor="FFF3CD")
                        cell.font = Font(name="Calibri", bold=True, color="856404", size=9)
            _aw(ws_gp); ws_gp.freeze_panes = "A2"

        # GARCH varianza condicional y proyección: formato numérico
        for sh_name, nf in [("GARCH_Var_Cond", "0.0000%"),
                              ("GARCH_Proyeccion", "0.00%")]:
            if sh_name in wb.sheetnames:
                _fmt(wb[sh_name], nf)

    print(f"\n  Excel guardado: {os.path.basename(path_excel)}")
    return path_excel


# =============================================================================
# MAIN
# =============================================================================
if __name__ == "__main__":

    try:
        import yfinance, pandas, numpy, matplotlib, seaborn, scipy, statsmodels, openpyxl
    except ImportError as e:
        print(f"!! Libreria faltante: {e}")
        print("   pip install yfinance pandas numpy matplotlib seaborn scipy statsmodels openpyxl arch")
        raise SystemExit(1)

    if not ARCH_OK:
        print("ADVERTENCIA: paquete 'arch' no encontrado.")
        print("Phillips-Perron y GARCH seran omitidos.")
        print("Instala con:  pip install arch\n")

    tickers, start, end, nombres = solicitar_configuracion()
    palette = construir_paleta(tickers)

    print("\n[0b] CONFIGURACION DE SALIDA")
    print("-" * 50)
    output_dir, base_name = obtener_destino()

    # ── Pipeline ──────────────────────────────────────────────────────────
    precios       = descargar_datos(tickers, start, end)
    precios_clean = control_calidad(precios)
    retornos      = calcular_retornos(precios_clean)
    tabla_tests   = tests_estacionariedad(precios_clean, retornos)
    tabla_arch    = test_arch_lm(retornos, lags_list=[5, 10, 20])
    tabla_stats   = estadisticos_descriptivos(retornos, nombres)
    garch_res     = estimar_garch(retornos)

    print("\n[8/9] GENERANDO GRAFICOS")
    print("-" * 50)
    graficos_serie_niveles(precios_clean,       palette, nombres, output_dir, base_name)
    graficos_serie_retornos(retornos,           palette, nombres, output_dir, base_name)
    grafico_distribuciones(retornos,            palette, nombres, output_dir, base_name)
    grafico_heatmap_correlacion(retornos,       output_dir, base_name)
    graficos_garch_volatilidad(retornos, garch_res, palette, nombres, output_dir, base_name)
    grafico_proyeccion_garch(garch_res,         palette, output_dir, base_name)

    print("\n[9/9] EXPORTANDO A EXCEL")
    print("-" * 50)
    exportar_excel(precios_clean, retornos, tabla_stats, tabla_tests,
                   tabla_arch, garch_res, output_dir, base_name)

    n = len(tickers)
    print("\n" + "=" * 60)
    print("  ANALISIS COMPLETADO")
    print(f"  Archivos en   : {output_dir}")
    print(f"  PNG niveles   : {n}   (uno por activo)")
    print(f"  PNG retornos  : {n}   (uno por activo)")
    print(f"  PNG GARCH     : {n}   (retornos + volatilidad condicional)")
    print(f"  PNG extras    : 3   (distribuciones + correlacion + proyeccion)")
    print(f"  Total PNG     : {3*n + 3}   a 300 dpi")
    print(f"  Excel XLSX    : 1 archivo, {'8' if ARCH_OK else '5'} hojas")
    print("=" * 60)